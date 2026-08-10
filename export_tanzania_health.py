#!/usr/bin/env python3
"""
export_tanzania_health.py
==========================
Tengeneza data kamili ya VITUO VYA AFYA Tanzania (hospitali, vituo vya afya,
zahanati n.k.) kutoka HFR Portal (Ministry of Health) — zilizopangwa kwa
mikoa yote na wilaya zote, kama tulivyofanya shule.

Chanzo: https://hfrportal.moh.go.tz (Health Facility Registry, MOH Tanzania)

Matokeo:
    tanzania_health_data/
        README.md
        Tanzania_Vituo_vya_Afya.xlsx      (sheets: Mikoa, Wilaya, Vituo vya Afya)
        csv/facilities.csv  |  csv/summary_regions.csv  |  csv/summary_districts.csv
        json/facilities.json  |  json/summary.json
"""

import csv
import html
import json
import os
import re
import sys
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(BASE_DIR, "tanzania_health_data")

# ── Rangi za bendera ya Tanzania ─────────────────────────────────────────────
GREEN = "FF1EB53A"
GOLD  = "FFFCD116"
BLUE  = "FF00A3DD"
BLACK = "FF000000"
WHITE = "FFFFFFFF"
LIGHT_GREY = "FFF2F2F2"
DARK_GREY  = "FF333333"
RED   = "FFD32F2F"   # kwa afya (msalaba mwekundu accent)

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _border(color="FFCCCCCC"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def _green_header(cell, size=10):
    cell.font = Font(color=WHITE, bold=True, name='Calibri', size=size)
    cell.fill = _fill(GREEN)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = _border(GOLD)

def _auto_width(ws, min_w=8, max_w=45):
    for col in ws.columns:
        max_len = min_w
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[letter].width = min(max_len + 3, max_w)

# ── Load data ────────────────────────────────────────────────────────────────
def load_fixture(path):
    with open(path, encoding='utf-8') as f:
        return json.load(f)

def load_regions_districts():
    """Regions & districts from our fixtures (188 wilaya, 26 mikoa)."""
    regions = load_fixture(os.path.join(BASE_DIR, 'regions.json'))
    districts = load_fixture(os.path.join(BASE_DIR, 'districts.json'))
    region_by_id = {r['pk']: r['fields']['name'] for r in regions}
    district_rows = []
    for d in districts:
        f = d['fields']
        district_rows.append({
            'name': f['name'],
            'region': region_by_id.get(f.get('region'), ''),
            'district_id': d['pk'],
        })
    return regions, district_rows, region_by_id

def normalize(s):
    """Normalize council/district names for matching."""
    s = html.unescape(s or '')
    s = s.replace('&#039;', "'").replace('&amp;', '&')
    s = s.strip().lower()
    # remove council suffix variants
    s = re.sub(r'\b(mc|dc|tc|cc|cc)\b', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def load_facilities():
    """Load scraped HFR facilities (fallback to partial if full missing)."""
    full = os.path.join('/tmp/health_data', 'hfr_facilities.json')
    partial = os.path.join('/tmp/health_data', 'hfr_partial.json')
    for path in (full, partial):
        if os.path.exists(path):
            with open(path, encoding='utf-8') as f:
                data = json.load(f)
            if data:
                return data
    raise SystemExit("Hakuna data ya HFR — endesha scrape_hfr.py kwanza.")

def classify_type(raw):
    """Group HFR facility types into a simple Tanzania category."""
    t = (raw or '').strip().lower()
    if 'hospital' in t or 'regional referral' in t or 'national' in t:
        if 'district' in t:
            return 'Hospitali ya Wilaya'
        if 'regional' in t:
            return 'Hospitali ya Mkoa (Rufaa)'
        if 'national' in t or 'referral' in t:
            return 'Hospitali ya Taifa (Rufaa)'
        return 'Hospitali'
    if 'health center' in t:
        return 'Kituo cha Afya'
    if 'dispensary laboratory' in t or 'laboratory' in t:
        return 'Maabara (Laboratory)'
    if 'dispensary' in t:
        return 'Zahanati (Dispensary)'
    if 'clinic' in t or 'polyclinic' in t or 'polyclinic' in t:
        return 'Kliniki'
    if 'maternity' in t or 'nursing home' in t:
        return 'Kituo cha Uzazi'
    if 'vaccine' in t or 'store' in t:
        return 'Hifadhi ya Chanjo'
    if 'dental' in t:
        return 'Kliniki ya Meno'
    if 'optometry' in t or 'eye' in t:
        return 'Kliniki ya Macho'
    if 'mobile' in t:
        return 'Kituo cha Mkononi (Mobile)'
    return 'Kingine'

def normalize_type_key(raw):
    t = (raw or '').strip().lower()
    if 'hospital' in t:
        return 'hospital'
    if 'health center' in t:
        return 'health_center'
    if 'dispensary' in t or 'dispensary laboratory' in t:
        return 'dispensary'
    return 'other'

# ── Build enriched facility list ─────────────────────────────────────────────
def build_enriched(facilities, district_rows, region_by_id):
    """Match each facility to our 188 districts / 26 regions."""
    # index districts by normalized name
    dist_by_norm = {}
    for d in district_rows:
        dist_by_norm.setdefault(normalize(d['name']), d)

    # fallback: match by region only (for unmatched councils)
    # build region name -> list of districts
    districts_by_region = OrderedDict()
    for d in district_rows:
        districts_by_region.setdefault(d['region'], []).append(d)

    enriched = []
    unmatched = OrderedDict()

    for f in facilities:
        raw_council = (f.get('council') or '').strip()
        raw_region = (f.get('region') or '').strip().replace(' Region', '').strip()
        norm = normalize(raw_council)

        rec = dict(f)
        rec['type_category'] = classify_type(f.get('type'))
        rec['type_key'] = normalize_type_key(f.get('type'))
        rec['region_normalized'] = raw_region
        rec['district_matched'] = ''
        rec['region_matched'] = ''

        d = dist_by_norm.get(norm)
        if d:
            rec['district_matched'] = d['name']
            rec['region_matched'] = d['region']
        else:
            unmatched.setdefault(raw_council or '(bila wilaya)', 0)
            unmatched[raw_council or '(bila wilaya)'] += 1
            # try to infer region from our districts
            for rn, ds in districts_by_region.items():
                if rn.strip().lower() == raw_region.strip().lower() and ds:
                    rec['district_matched'] = ds[0]['name']
                    rec['region_matched'] = rn
                    break
            if not rec['region_matched']:
                rec['region_matched'] = raw_region or '(bila mkoa)'
        enriched.append(rec)

    return enriched, unmatched

# ── Build Excel ──────────────────────────────────────────────────────────────
def build_workbook(enriched, district_rows):
    wb = openpyxl.Workbook()

    # ═══ SHEET 1: MIKOA ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Mikoa"
    headers = ["NAFASI", "MKOA", "IDADI YA WILAYA", "VITUO VYA AFYA", "HOSPITALI", "VITUO VYA AFYA", "ZAHANATI"]
    headers = ["NAFASI", "MKOA", "IDADI YA WILAYA", "IDADI YA VITUO", "HOSPITALI", "VITUO VYA AFYA", "ZAHANATI"]
    for col, h in enumerate(headers, 1):
        _green_header(ws.cell(row=1, column=col, value=h))
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # aggregate
    agg = OrderedDict()
    for rec in enriched:
        rn = rec.get('region_matched') or '(bila mkoa)'
        agg.setdefault(rn, {'n': 0, 'hospital': 0, 'health_center': 0, 'dispensary': 0})
        a = agg[rn]
        a['n'] += 1
        a[rec.get('type_key', 'other')] = a.get(rec.get('type_key', 'other'), 0) + 1

    # district counts per region
    dist_counts = {}
    for d in district_rows:
        dist_counts[d['region']] = dist_counts.get(d['region'], 0) + 1

    row = 2
    for i, (rn, a) in enumerate(sorted(agg.items()), 1):
        vals = [i, rn, dist_counts.get(rn, 0), a['n'], a.get('hospital', 0), a.get('health_center', 0), a.get('dispensary', 0)]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = Font(name='Calibri', size=10)
            cell.border = _border()
            if col != 2:
                cell.alignment = Alignment(horizontal='center')
            if row % 2 == 0:
                cell.fill = _fill(LIGHT_GREY)
        ws.cell(row=row, column=2).font = Font(name='Calibri', size=10, bold=True)
        row += 1
    _auto_width(ws)

    # ═══ SHEET 2: WILAYA ═════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Wilaya")
    headers2 = ["MKOA", "WILAYA", "IDADI YA VITUO", "HOSPITALI", "VITUO VYA AFYA", "ZAHANATI"]
    for col, h in enumerate(headers2, 1):
        _green_header(ws2.cell(row=1, column=col, value=h))
    ws2.row_dimensions[1].height = 22
    ws2.freeze_panes = "A2"

    dist_agg = OrderedDict()
    for rec in enriched:
        dm = rec.get('district_matched') or '(bila wilaya)'
        rm = rec.get('region_matched') or ''
        key = (rm, dm)
        dist_agg.setdefault(key, {'n': 0, 'hospital': 0, 'health_center': 0, 'dispensary': 0})
        a = dist_agg[key]
        a['n'] += 1
        a[rec.get('type_key', 'other')] = a.get(rec.get('type_key', 'other'), 0) + 1

    row = 2
    for (rm, dm), a in sorted(dist_agg.items()):
        vals = [rm, dm, a['n'], a.get('hospital', 0), a.get('health_center', 0), a.get('dispensary', 0)]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=v)
            cell.font = Font(name='Calibri', size=10)
            cell.border = _border()
            if col != 2:
                cell.alignment = Alignment(horizontal='center')
            if row % 2 == 0:
                cell.fill = _fill(LIGHT_GREY)
        ws2.cell(row=row, column=2).font = Font(name='Calibri', size=10, bold=True)
        row += 1
    _auto_width(ws2)

    # ═══ SHEET 3: VITUO VYA AFYA ═════════════════════════════════════════════
    ws3 = wb.create_sheet("Vituo vya Afya")
    headers3 = ["MKOA", "WILAYA", "CODE", "JINA LA KITUO", "AINA YA KITUO", "KAWAIDA (TYPE)", "UMILIKI (CATEGORY)", "UMILIKI (AUTHORITY)", "HALI (STATUS)"]
    for col, h in enumerate(headers3, 1):
        _green_header(ws3.cell(row=1, column=col, value=h))
    ws3.row_dimensions[1].height = 22
    ws3.freeze_panes = "A2"

    row = 2
    for rec in sorted(enriched, key=lambda x: (x.get('region_matched') or '', x.get('district_matched') or '', x.get('name') or '')):
        vals = [
            rec.get('region_matched', ''),
            rec.get('district_matched', ''),
            rec.get('code', ''),
            rec.get('name', ''),
            rec.get('type', ''),
            rec.get('type_category', ''),
            rec.get('ownership_category', ''),
            rec.get('ownership_authority', ''),
            rec.get('status', ''),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=col, value=v)
            cell.font = Font(name='Calibri', size=9)
            cell.border = _border()
            if row % 2 == 0:
                cell.fill = _fill(LIGHT_GREY)
        ws3.cell(row=row, column=4).font = Font(name='Calibri', size=9, bold=True)
        row += 1
    _auto_width(ws3)
    ws3.column_dimensions['D'].width = 45

    return wb

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    print("Loading regions & districts...")
    regions, district_rows, region_by_id = load_regions_districts()
    print(f"  Mikoa: {len(regions)} | Wilaya zetu: {len(district_rows)}")

    print("Loading HFR facilities...")
    facilities = load_facilities()
    print(f"  Vituo vya HFR: {len(facilities)}")

    print("Matching councils -> wilaya zetu...")
    enriched, unmatched = build_enriched(facilities, district_rows, region_by_id)
    matched = sum(1 for r in enriched if r['district_matched'])
    print(f"  Vituo vilivyofananishwa na wilaya: {matched}/{len(enriched)}")
    if unmatched:
        print(f"  Councils ambazo hazikufananishwa ({len(unmatched)}):")
        for name, cnt in list(unmatched.items())[:15]:
            print(f"    - {name}: {cnt}")

    print("Building Excel...")
    wb = build_workbook(enriched, district_rows)
    os.makedirs(OUT_DIR, exist_ok=True)
    os.makedirs(os.path.join(OUT_DIR, 'csv'), exist_ok=True)
    os.makedirs(os.path.join(OUT_DIR, 'json'), exist_ok=True)
    xlsx = os.path.join(OUT_DIR, "Tanzania_Vituo_vya_Afya.xlsx")
    wb.save(xlsx)
    print(f"  ✅ {xlsx}")

    print("Writing CSV & JSON...")
    # facilities CSV (clean fields only)
    fac_rows = []
    for r in enriched:
        fac_rows.append({
            'region': r.get('region_matched', ''),
            'district': r.get('district_matched', ''),
            'code': r.get('code', ''),
            'name': r.get('name', ''),
            'type': r.get('type', ''),
            'type_category': r.get('type_category', ''),
            'ownership_category': r.get('ownership_category', ''),
            'ownership_authority': r.get('ownership_authority', ''),
            'status': r.get('status', ''),
        })
    with open(os.path.join(OUT_DIR, 'csv', 'facilities.csv'), 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=list(fac_rows[0].keys()))
        w.writeheader()
        w.writerows(fac_rows)

    # summary CSVs
    def write_summary_csv(fname, rows, fieldnames):
        with open(os.path.join(OUT_DIR, 'csv', fname), 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)

    # region summary
    region_rows = []
    agg = OrderedDict()
    for r in enriched:
        rn = r.get('region_matched') or '(bila mkoa)'
        agg.setdefault(rn, {'n': 0, 'hospital': 0, 'health_center': 0, 'dispensary': 0})
        a = agg[rn]
        a['n'] += 1
        a[r.get('type_key', 'other')] = a.get(r.get('type_key', 'other'), 0) + 1
    for rn, a in sorted(agg.items()):
        region_rows.append({'region': rn, 'facilities': a['n'], 'hospitals': a.get('hospital', 0),
                            'health_centers': a.get('health_center', 0), 'dispensaries': a.get('dispensary', 0)})
    write_summary_csv('summary_regions.csv', region_rows, ['region', 'facilities', 'hospitals', 'health_centers', 'dispensaries'])

    # district summary
    district_rows_out = []
    dist_agg = OrderedDict()
    for r in enriched:
        dm = r.get('district_matched') or '(bila wilaya)'
        rm = r.get('region_matched') or ''
        key = (rm, dm)
        dist_agg.setdefault(key, {'n': 0, 'hospital': 0, 'health_center': 0, 'dispensary': 0})
        a = dist_agg[key]
        a['n'] += 1
        a[r.get('type_key', 'other')] = a.get(r.get('type_key', 'other'), 0) + 1
    for (rm, dm), a in sorted(dist_agg.items()):
        district_rows_out.append({'region': rm, 'district': dm, 'facilities': a['n'],
                                  'hospitals': a.get('hospital', 0), 'health_centers': a.get('health_center', 0),
                                  'dispensaries': a.get('dispensary', 0)})
    write_summary_csv('summary_districts.csv', district_rows_out, ['region', 'district', 'facilities', 'hospitals', 'health_centers', 'dispensaries'])

    # JSON
    with open(os.path.join(OUT_DIR, 'json', 'facilities.json'), 'w', encoding='utf-8') as f:
        json.dump(fac_rows, f, ensure_ascii=False, indent=1)
    summary = {
        'total_facilities': len(fac_rows),
        'source': 'HFR Portal (hfrportal.moh.go.tz) - Ministry of Health, Tanzania',
        'matched_districts': len(set((r['region'], r['district']) for r in fac_rows)),
        'generated': '2026-08-04',
    }
    with open(os.path.join(OUT_DIR, 'json', 'summary.json'), 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=1)

    print(f"\n✅ DONE: {len(fac_rows)} vituo vya afya -> {OUT_DIR}/")
    return 0

if __name__ == '__main__':
    sys.exit(main())
