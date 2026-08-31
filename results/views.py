import logging
import os
import re
from pathlib import Path

logger = logging.getLogger(__name__)

import pandas as pd

from django.conf import settings
from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.db.models import Case, Count, IntegerField, Q, When
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET

from django.core.exceptions import ValidationError

from .forms import ExamUploadForm, TeacherSelfSubjectsForm
from .models import Exam, ExamResult, FormStudent, PersonalUpload, PersonalUploadResult, PrintSubmission, ProcessedResult, School, SchoolSubject, Student, Subject, SubjectSubmission, TeacherFormAssignment
from .permissions import academic_required, printing_secretary_required, results_login_required as login_required, teacher_required
from .services.excel_export_service import generate_professional_excel_response, generate_results_excel_response
from .services.pdf_export_service import (
    generate_bulk_student_results_pdf_response,
    generate_results_pdf_response,
    generate_student_result_pdf_response,
)
from .services.subject_pdf_service import (
    GRADE_KEYS_OLEVEL,
    generate_personal_pdf_response,
    generate_subject_pdf_response,
)
from .services.results_analytics import compute_subject_stats, generate_recommendations
from .services.upload_processing_service import (
    UploadProcessingError,
    process_uploaded_results,
    recompute_processed_results_for_exam,
)
from .utils import get_grade, get_grade_for_form, normalize_gender, parse_name_score_sheet, parse_score, safe_get_or_create_subject

_EXAM_TYPE_CHOICES = Exam.EXAM_TYPE_CHOICES


def _get_exam_or_404(exam_id, user):
    """Safe exam lookup — tolerates NULL school FK and school mismatches.

    1. Find by ID.
    2. If user has a school and exam.school matches → OK.
    3. If exam.school is NULL or mismatched → check teacher owns subjects on this exam.
    4. Otherwise → 404.
    """
    exam = Exam.objects.filter(id=exam_id).first()
    if exam is None:
        raise Http404("Mtihani haujapatikana.")
    user_school = getattr(user, 'school', None)
    # Fast path: schools match
    if user_school and exam.school_id == user_school.id:
        return exam
    # exam.school may be NULL or wrong — check teacher access
    has_access = (
        exam.subject_submissions.filter(
            subject__in=user.subjects.all()
        ).exists()
        or exam.subject_submissions.filter(submitted_by_user=user).exists()
    )
    if has_access:
        return exam
    raise Http404("Mtihani haujapatikana kwenye shule yako.")


COMMON_SUBJECTS = [
    'Mathematics', 'English', 'Kiswahili', 'Biology', 'Chemistry',
    'Physics', 'History', 'Geography', 'Civics', 'Computer Studies',
    'Agriculture', 'Business Studies', 'CRE', 'IRE', 'Fine Art',
    'Music', 'Physical Education', 'Further Mathematics',
]


@login_required
def home(request):
    exams = Exam.objects.filter(school=request.user.school).order_by('-year', 'name').annotate(
        submitted_count=Count(
            'subject_submissions',
            filter=Q(subject_submissions__status=SubjectSubmission.STATUS_SUBMITTED),
        ),
        total_submissions=Count('subject_submissions'),
    )

    exams_list = list(exams)

    return render(
        request,
        'results/home.html',
        {
            'exams': exams_list,
            'exam_count': len(exams_list),
            'latest_exam': exams_list[0] if exams_list else None,
        },
    )


@academic_required
def upload_results(request):
    school = request.user.school
    no_exams = not Exam.objects.filter(school=school).exists()

    if request.method == 'POST':
        action = request.POST.get('action', 'upload')

        if action == 'create_exam':
            import json as _json
            name = request.POST.get('exam_name', '').strip()
            year = request.POST.get('exam_year', '2026')
            form_level = request.POST.get('exam_form', '4')
            stream = request.POST.get('exam_stream', '').strip()
            exam_type = request.POST.get('exam_type_new', 'TERMINAL')
            subjects_raw = request.POST.get('subjects', '[]')
            try:
                subject_names = [s.strip() for s in _json.loads(subjects_raw) if str(s).strip()]
            except Exception:
                subject_names = []

            if name:
                exam, _ = Exam.objects.get_or_create(
                    name=name, year=int(year), form=int(form_level), stream=stream,
                    exam_type=exam_type, school=school,
                    defaults={'school_name': school.name if school else ''},
                )

                # Create subjects + SubjectSubmission (PENDING) for each
                for sname in subject_names:
                    subject = safe_get_or_create_subject(sname)
                    SubjectSubmission.objects.get_or_create(exam=exam, subject=subject)

                # Redirect to exam overview — main hub for teachers
                return redirect(reverse('exam_overview', args=[exam.id]))

        form = ExamUploadForm(request.POST, request.FILES, school=school)
        if form.is_valid():
            exam = form.cleaned_data['exam']
            file = form.cleaned_data['file']
            try:
                process_uploaded_results(exam=exam, uploaded_file=file)
                messages.success(request, f"Matokeo yamepakiwa: {exam.name}")
                download_url = reverse('generate_results_pdf', args=[exam.id])
                return render(request, 'results/upload.html', {
                    'form': ExamUploadForm(school=school),
                    'download_url': download_url,
                    'no_exams': False,
                    'exam_type_choices': _EXAM_TYPE_CHOICES,
                })
            except UploadProcessingError as error:
                messages.error(request, str(error))
                return redirect(request.path)
            except Exception as e:
                messages.error(request, f"Hitilafu: {str(e)}")
                return redirect(request.path)
    else:
        pass

    return render(request, 'results/upload.html', {
        'exam_type_choices': _EXAM_TYPE_CHOICES,
        'common_subjects': COMMON_SUBJECTS,
        'form': ExamUploadForm(school=school),
    })


@login_required
def filter_exams(request):
    exam_type = request.GET.get('exam_type')
    exams = Exam.objects.filter(school=request.user.school)
    if exam_type:
        exams = exams.filter(exam_type=exam_type)

    options_html = render_to_string('results/exam_options.html', {'exams': exams})
    return HttpResponse(options_html)


@academic_required
def generate_results_pdf(request, exam_id):
    exam = _get_exam_or_404(exam_id, request.user)
    recompute_processed_results_for_exam(exam)
    return generate_results_pdf_response(exam)


@academic_required
def generate_bulk_student_results_pdf(request, exam_id):
    """All students of this exam, each on their own result-slip page(s),
    merged into one downloadable PDF — the bulk counterpart to the
    single-student download at /shule/matokeo/<token>/pdf/. Restricted to
    the Academic Officer since it exposes every student's result in one
    file (the single-student one is safe to be public because a parent
    only has their own child's share token)."""
    exam = _get_exam_or_404(exam_id, request.user)
    recompute_processed_results_for_exam(exam)
    return generate_bulk_student_results_pdf_response(exam)


@academic_required
def export_results_excel(request, exam_id):
    exam = _get_exam_or_404(exam_id, request.user)
    recompute_processed_results_for_exam(exam)
    return generate_results_excel_response(exam)


# ── Printing Secretary (PS) handoff ────────────────────────────────────────
# Academic Officer submits the exam's results PDF to the school's Printing
# Secretary; the secretary sees it on their own dashboard and prints it.
# The PDF is never stored — it's regenerated on demand from the exam, same
# as everywhere else in this app; PrintSubmission just tracks the handoff.

@academic_required
@require_POST
def submit_exam_to_ps(request, exam_id):
    exam = _get_exam_or_404(exam_id, request.user)
    submission, created = PrintSubmission.objects.get_or_create(
        exam=exam,
        defaults={'school': exam.school, 'submitted_by': request.user},
    )
    if not created:
        submission.status = PrintSubmission.STATUS_PENDING
        submission.school = exam.school
        submission.submitted_by = request.user
        submission.submitted_at = timezone.now()
        submission.printed_by = None
        submission.printed_at = None
        submission.save()
    messages.success(request, "Matokeo yametumwa kwa Katibu wa Uchapishaji (PS).")
    return redirect('exam_overview', exam_id=exam.id)


@printing_secretary_required
def printing_secretary_dashboard(request):
    submissions = PrintSubmission.objects.filter(
        school=request.user.school
    ).select_related('exam', 'submitted_by', 'printed_by')
    pending = [s for s in submissions if s.status == PrintSubmission.STATUS_PENDING]
    printed = [s for s in submissions if s.status == PrintSubmission.STATUS_PRINTED]

    # Timetable submissions
    from .models import TimetablePrintSubmission
    timetable_subs = TimetablePrintSubmission.objects.filter(
        school=request.user.school
    ).select_related('submitted_by', 'printed_by')
    timetable_pending = [s for s in timetable_subs if s.status == TimetablePrintSubmission.STATUS_PENDING]
    timetable_printed = [s for s in timetable_subs if s.status == TimetablePrintSubmission.STATUS_PRINTED]

    return render(request, 'results/printing_secretary_dashboard.html', {
        'pending_submissions': pending,
        'printed_submissions': printed,
        'timetable_pending': timetable_pending,
        'timetable_printed': timetable_printed,
    })


@printing_secretary_required
def ps_pdf_inline(request, submission_id):
    """Serves the results PDF inline (not as an attachment) so the PS
    print page can embed it in an <iframe> and trigger the browser's
    print dialog on it."""
    submission = get_object_or_404(PrintSubmission, pk=submission_id, school=request.user.school)
    response = generate_results_pdf_response(submission.exam)
    filename = f"Matokeo_{submission.exam.id}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@printing_secretary_required
@require_POST
def ps_print_view(request, submission_id):
    """Records the print action and hands back a page that auto-opens the
    browser print dialog on the embedded PDF."""
    submission = get_object_or_404(PrintSubmission, pk=submission_id, school=request.user.school)
    submission.status = PrintSubmission.STATUS_PRINTED
    submission.printed_by = request.user
    submission.printed_at = timezone.now()
    submission.save()
    return render(request, 'results/ps_print.html', {
        'submission': submission,
        'pdf_url': reverse('ps_pdf_inline', args=[submission.id]),
    })


# ── Public Results Portal — Search Page (NECTA-style, no login required) ───

def public_results_search(request):
    """Public search portal — no login required.
    Parents can search for their child's results by name.
    Like NECTA's online results portal."""
    query = request.GET.get('q', '').strip()
    selected_form = request.GET.get('form', '').strip()
    selected_year = request.GET.get('year', '').strip()

    results = []
    has_searched = bool(query)
    lang = 'sw'  # Default to Swahili for public page

    # Detect language from browser
    accept_lang = request.META.get('HTTP_ACCEPT_LANGUAGE', '')
    if 'en' in accept_lang and 'sw' not in accept_lang:
        lang = 'en'

    form_choices = range(1, 7)
    current_year = timezone.now().year
    year_choices = range(current_year, current_year - 5, -1)

    if has_searched:
        # Build query for ProcessedResult, filtering by student name
        from django.db.models import Q

        qs = ProcessedResult.objects.select_related(
            'student', 'exam', 'exam__school'
        ).all()

        # Filter by form
        if selected_form:
            try:
                qs = qs.filter(exam__form=int(selected_form))
            except ValueError:
                pass

        # Filter by year
        if selected_year:
            try:
                qs = qs.filter(exam__year=int(selected_year))
            except ValueError:
                pass

        # Search by name — split query into parts and match across name fields
        name_parts = query.split()
        name_filter = Q()
        for part in name_parts:
            name_filter &= (
                Q(student__first_name__icontains=part) |
                Q(student__middle_name__icontains=part) |
                Q(student__last_name__icontains=part)
            )
        qs = qs.filter(name_filter)

        # Order by exam year descending, then position
        qs = qs.order_by('-exam__year', 'position')

        # Limit to recent results (max 50 to avoid overload)
        qs = qs[:50]

        for r in qs:
            st = r.student
            name = ' '.join(p for p in [st.first_name, st.middle_name or '', st.last_name] if p)
            school_name = r.exam.school_name or (r.exam.school.name if r.exam.school else '')
            results.append({
                'student_name': name,
                'school_name': school_name,
                'exam_type': r.exam.get_exam_type_display(),
                'form': r.exam.form,
                'year': r.exam.year,
                'division': r.division,
                'position': r.position,
                'share_token': str(r.share_token),
            })

    return render(request, 'results/student_results_search.html', {
        'query': query,
        'selected_form': selected_form,
        'selected_year': selected_year,
        'results': results,
        'has_searched': has_searched,
        'form_choices': form_choices,
        'year_choices': year_choices,
        'lang': lang,
    })


# ── Public Results Lookup (NECTA-style, no login required) ───────────────────

def student_result_public(request, token):
    """Public view — no login required. Look up a student's result by share token.
    Returns a clean, professional results page like NECTA's online portal."""
    result = get_object_or_404(
        ProcessedResult.objects.select_related('exam__school', 'student'), share_token=token
    )
    exam = result.exam
    student = result.student

    # Get all subjects, scores and absent markers for this exam+student
    # Only subjects the student is enrolled in (has an ExamResult entry)
    student_results = ExamResult.objects.filter(exam=exam, student=student)
    enrolled_subject_ids = set()
    scores = {}
    absent_subjects = set()
    for er in student_results:
        enrolled_subject_ids.add(er.subject_id)
        if er.is_absent:
            absent_subjects.add(er.subject_id)
        else:
            scores[er.subject_id] = er.score

    subjects = list(
        Subject.objects.filter(id__in=enrolled_subject_ids).order_by('name')
    )

    # Prepare row data for each subject
    subject_rows = []
    for subj in subjects:
        if subj.id in absent_subjects:
            # Student studies this subject but was absent
            subject_rows.append({
                'subject': subj.name,
                'score': None,
                'grade': 'X',
                'is_absent': True,
            })
        else:
            score = scores.get(subj.id)
            if score is not None:
                subject_rows.append({
                    'subject': subj.name,
                    'score': score,
                    'grade': get_grade_for_form(score, exam.form),
                    'is_absent': False,
                })

    student_name = ' '.join(p for p in [student.first_name, student.middle_name or '', student.last_name] if p)
    location = ''
    if exam.school:
        parts = []
        if exam.school.district:
            parts.append(exam.school.district)
        if exam.school.region:
            parts.append(exam.school.region)
        location = ', '.join(parts)

    division_label = dict(ProcessedResult.DIVISION_CHOICES).get(result.division, result.division)

    from .services.subject_pdf_service import get_grade_keys_for_form
    grade_key = get_grade_keys_for_form(exam.form)

    return render(request, 'results/student_result_public.html', {
        'result': result,
        'exam': exam,
        'student': student,
        'student_name': student_name,
        'subject_rows': subject_rows,
        'division_label': division_label,
        'grade_key': grade_key,
        'location': location,
        'school_name': exam.school_name or (exam.school.name if exam.school else ''),
        'total_students': ProcessedResult.objects.filter(exam=exam).count(),
    })


def student_result_pdf(request, token):
    """Downloadable version of the public results page — no login required,
    same share token, so a parent can save/print a copy to take home."""
    result = get_object_or_404(ProcessedResult, share_token=token)
    return generate_student_result_pdf_response(result)


# ── Shareable Links Management (academic only) ───────────────────────────────

@academic_required
def exam_share_links(request, exam_id):
    """Academic officer views all shareable links for an exam's students.
    Each link can be copied and shared with parents for online results lookup."""
    exam = _get_exam_or_404(exam_id, request.user)
    results = list(
        ProcessedResult.objects.filter(exam=exam)
        .select_related('student')
        .order_by('position')
    )

    # Build the base URL from request
    base_url = f"{request.scheme}://{request.get_host()}"

    students_links = []
    for r in results:
        st = r.student
        name = ' '.join(p for p in [st.first_name, st.middle_name or '', st.last_name] if p)
        full_url = f"{base_url}/shule/matokeo/{r.share_token}/"
        students_links.append({
            'position': r.position,
            'name': name,
            'division': r.division,
            'token': str(r.share_token),
            'full_url': full_url,
        })

    school_name = exam.school_name or (exam.school.name if exam.school else '')
    etype = exam.get_exam_type_display()

    return render(request, 'results/exam_share_links.html', {
        'exam': exam,
        'school_name': school_name,
        'etype': etype,
        'students_links': students_links,
        'total_students': len(students_links),
        'base_url': base_url,
    })


# ── Exam Overview Dashboard ───────────────────────────────────────────────────

@login_required
def exam_overview(request, exam_id):
    exam = _get_exam_or_404(exam_id, request.user)
    is_academic = getattr(request.user, 'is_academic', False)

    # Get all subjects that have results for this exam + known subjects from submissions
    subjects_with_results = list(
        Subject.objects.filter(examresult__exam=exam).distinct().order_by('name')
    )
    # Also include subjects from existing submissions
    submitted_subject_ids = set(
        exam.subject_submissions.values_list('subject_id', flat=True)
    )
    extra_subjects = Subject.objects.filter(id__in=submitted_subject_ids).exclude(
        id__in=[s.id for s in subjects_with_results]
    )
    all_subjects = list(subjects_with_results) + list(extra_subjects)
    all_subjects.sort(key=lambda s: s.name)

    # Build submission map
    submission_map = {
        sub.subject_id: sub
        for sub in exam.subject_submissions.select_related('subject').all()
    }

    # Build per-subject context
    subjects_ctx = []
    submitted_count = 0
    approved_count = 0
    for subject in all_subjects:
        submission = submission_map.get(subject.id)
        is_submitted = submission and submission.status == SubjectSubmission.STATUS_SUBMITTED
        is_approved = submission and submission.status == SubjectSubmission.STATUS_APPROVED
        is_returned = submission and submission.status == SubjectSubmission.STATUS_RETURNED
        if is_submitted or is_approved:
            submitted_count += 1
        if is_approved:
            approved_count += 1
        subjects_ctx.append({
            'subject': subject,
            'submission': submission,
            'is_submitted': is_submitted,
            'is_approved': is_approved,
            'is_returned': is_returned,
            'marks_url': reverse('marks_entry') + f'?exam={exam.id}&subject={subject.id}',
            'upload_url': reverse('subject_upload', args=[exam.id, subject.id]),
            'pdf_url': reverse('subject_pdf', args=[exam.id, subject.id]) if (is_submitted or is_approved) else None,
            'approve_url': reverse('approve_subject', args=[exam.id, subject.id]) if (is_submitted and is_academic) else None,
            'return_url': reverse('return_submission', args=[exam.id, subject.id]) if (is_submitted and is_academic) else None,
        })

    total_subjects = len(all_subjects)
    all_submitted = submitted_count == total_subjects and total_subjects > 0
    all_approved = approved_count == total_subjects and total_subjects > 0
    # General school results are only ready once every subject for this exam
    # has been submitted by its teacher — partial results would misrepresent
    # each student's overall division/position.
    enough_to_finalize = all_submitted

    progress_pct = round(submitted_count / total_subjects * 100) if total_subjects else 0
    approval_pct = round(approved_count / total_subjects * 100) if total_subjects else 0

    return render(request, 'results/exam_overview.html', {
        'exam': exam,
        'subjects_ctx': subjects_ctx,
        'submitted_count': submitted_count,
        'approved_count': approved_count,
        'total_subjects': total_subjects,
        'all_submitted': all_submitted,
        'all_approved': all_approved,
        'enough_to_finalize': enough_to_finalize,
        'progress_pct': progress_pct,
        'approval_pct': approval_pct,
        'is_academic': is_academic,
        'finalize_url': reverse('finalize_exam', args=[exam.id]) if is_academic else None,
        'form_results_url': reverse('form_results', args=[exam.form]),
    })


# ── Subject Upload (CSV/Excel for one subject) ────────────────────────────────

@teacher_required
def subject_upload(request, exam_id, subject_id):
    exam = _get_exam_or_404(exam_id, request.user)
    subject = get_object_or_404(Subject, id=subject_id)
    if not request.user.subjects.filter(pk=subject.pk).exists():
        raise PermissionDenied("You are not assigned to teach this subject.")

    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            messages.error(request, "Hakuna faili lililochaguliwa.")
            return redirect(request.path)

        try:
            import io
            import pandas as pd

            uploaded_file.seek(0)
            file_name = uploaded_file.name.lower()

            # ── PDF roster upload (names only — no scores) ──
            if file_name.endswith('.pdf'):
                students_out = _bulk_save_students(_collect_roster_rows(uploaded_file, is_pdf=True))
                if not students_out:
                    raise UploadProcessingError(
                        "Hakuna wanafunzi waliopatikana kwenye PDF."
                    )
                # Return to marks_entry with these students pre-loaded
                # Store student IDs in session so marks_entry picks them up
                request.session['pending_roster_ids'] = [s['id'] for s in students_out]
                request.session['pending_roster_subject'] = subject.id
                request.session['pending_roster_exam'] = exam.id
                messages.success(
                    request,
                    f"Orodha ya wanafunzi {len(students_out)} imepakwa. Jaza alama na ubadilishe."
                )
                return redirect(
                    reverse('marks_entry') + f'?exam={exam.id}&subject={subject.id}'
                )

            # ── CSV / Excel upload (names + scores) ──
            if file_name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)

            # Normalize column names
            df.columns = [str(c).strip() for c in df.columns]

            # Expected columns: First Name, Last Name, Gender, Score
            # Try to find score column flexibly
            score_col = None
            for col in df.columns:
                if col.lower() in ('score', 'alama', 'marks', 'mark', 'result'):
                    score_col = col
                    break
            if score_col is None:
                # Try the last numeric column
                numeric_cols = [c for c in df.columns if pd.to_numeric(df[c], errors='coerce').notna().any()]
                if numeric_cols:
                    score_col = numeric_cols[-1]

            if score_col is None:
                raise UploadProcessingError("Hakuna safu ya alama iliyopatikana. Tumia jina kama 'Score' au 'Alama'.")

            parsed_rows = []
            for _, row in df.iterrows():
                first_name = str(row.get('First Name', row.get('first_name', ''))).strip()
                last_name = str(row.get('Last Name', row.get('last_name', ''))).strip()
                gender_raw = str(row.get('Gender', row.get('gender', 'M'))).strip()

                if not first_name or first_name in ('nan', 'None'):
                    continue

                raw_val = row.get(score_col)
                score_val = parse_score(raw_val)
                from .utils import is_absent_marker
                is_absent = is_absent_marker(raw_val)
                if score_val is None and not is_absent:
                    continue

                parsed_rows.append((first_name, last_name or 'Unknown', normalize_gender(gender_raw), score_val, is_absent))

            saved_count = 0
            if parsed_rows:
                # Bulk-create/get students using the same optimized path
                # as roster uploads — avoids N+1 get_or_create per row.
                name_tuples = [(fn, '', ln, g) for fn, ln, g, _, _ in parsed_rows]
                students_out = _bulk_save_students(name_tuples)
                student_id_map = {s['id']: s for s in students_out}
                # Also build (first, last) -> student for matching scores
                from .models import Student as _Stu
                all_first = {fn for fn, _, ln, _, _ in parsed_rows}
                all_last = {ln for fn, _, ln, _, _ in parsed_rows}
                student_map = {
                    (s.first_name, s.last_name): s
                    for s in _Stu.objects.filter(first_name__in=all_first, last_name__in=all_last)
                }

                exam_results = []
                for fn, ln, _, score_val, is_abs in parsed_rows:
                    student = student_map.get((fn, ln))
                    if not student:
                        continue
                    exam_results.append(ExamResult(
                        exam=exam, student=student, subject=subject,
                        score=score_val, is_absent=is_abs,
                    ))
                    saved_count += 1

                ExamResult.objects.bulk_create(
                    exam_results,
                    update_conflicts=True,
                    unique_fields=['exam', 'student', 'subject'],
                    update_fields=['score', 'is_absent'],
                )

            if saved_count == 0:
                messages.error(
                    request,
                    "Hakuna mwanafunzi aliyesomwa kwenye faili — angalia kama safu za 'First Name', "
                    "'Last Name' na 'Score' zipo na zina data sahihi. Hakuna kilichopakiwa."
                )
                return redirect(request.path)

            # Mark SubjectSubmission as SUBMITTED (update_or_create to avoid IntegrityError)
            SubjectSubmission.objects.update_or_create(
                exam=exam,
                subject=subject,
                defaults={
                    'status': SubjectSubmission.STATUS_SUBMITTED,
                    'method': 'UPLOAD',
                    'submitted_by': request.user.full_name or request.user.email,
                    'submitted_by_user': request.user,
                    'submitted_at': timezone.now(),
                    'student_count': saved_count,
                    'revision_number': SubjectSubmission.objects.filter(
                        exam=exam, subject=subject
                    ).values_list('revision_number', flat=True).first() or 1,
                },
            )

            # Recompute processed results
            recompute_processed_results_for_exam(exam)

            messages.success(
                request,
                f"Alama za {subject.name} zimepakiwa: wanafunzi {saved_count} wamesajiliwa."
            )
            return redirect(reverse('exam_overview', args=[exam.id]))

        except UploadProcessingError as e:
            messages.error(request, str(e))
            return redirect(request.path)
        except Exception as e:
            messages.error(request, f"Hitilafu ya faili: {e}")
            return redirect(request.path)

    return render(request, 'results/subject_upload.html', {
        'exam': exam,
        'subject': subject,
        'overview_url': reverse('exam_overview', args=[exam.id]),
    })


# ── Subject PDF Download ──────────────────────────────────────────────────────

@login_required
def subject_pdf(request, exam_id, subject_id):
    exam = _get_exam_or_404(exam_id, request.user)
    subject = get_object_or_404(Subject, id=subject_id)

    # Get teacher name from SubjectSubmission if available
    teacher_name = ''
    try:
        submission = SubjectSubmission.objects.get(exam=exam, subject=subject)
        teacher_name = submission.submitted_by or ''
    except SubjectSubmission.DoesNotExist:
        pass

    lang = request.session.get('ui_lang', 'en')
    return generate_subject_pdf_response(exam, subject, teacher_name=teacher_name, lang=lang)


# ── Subject Results Summary (in-app view: stats + recommendations) ───────────

@login_required
def subject_summary(request, exam_id, subject_id):
    """Same analysis as the PDF (distribution, gender, recommendations) but
    viewable directly in the app — teachers don't have to download anything
    to see how their subject did and what to do next."""
    exam = _get_exam_or_404(exam_id, request.user)
    subject = get_object_or_404(Subject, id=subject_id)

    teacher_name = ''
    try:
        submission = SubjectSubmission.objects.get(exam=exam, subject=subject)
        teacher_name = submission.submitted_by or ''
    except SubjectSubmission.DoesNotExist:
        pass

    results = list(
        ExamResult.objects.filter(exam=exam, subject=subject).select_related('student')
    )
    results.sort(key=lambda r: r.score, reverse=True)

    rows_data = []
    for pos, result in enumerate(results, 1):
        student = result.student
        full_name = ' '.join(part for part in [student.first_name, student.middle_name, student.last_name] if part)
        rows_data.append({
            'position': pos,
            'name': full_name,
            'score': result.score,
            'grade': get_grade_for_form(result.score, exam.form),
            'gender': student.gender,
        })

    lang = request.session.get('ui_lang', 'en')
    stats = compute_subject_stats(rows_data)
    recommendations = generate_recommendations(stats, subject_name=subject.name, lang=lang)
    from .services.subject_pdf_service import get_grade_keys_for_form
    grade_keys = get_grade_keys_for_form(exam.form)
    distribution = _build_distribution(stats, grade_keys)
    teacher_label = 'Mwalimu' if lang == 'sw' else 'Teacher'

    return render(request, 'results/results_summary.html', {
        'heading': subject.name,
        'meta_parts': [p for p in [exam.school_name, str(exam), f"{teacher_label}: {teacher_name}" if teacher_name else ''] if p],
        'rows_data': rows_data,
        'stats': stats,
        'recommendations': recommendations,
        'distribution': distribution,
        'pdf_url': reverse('subject_pdf', args=[exam.id, subject.id]),
        'back_url': reverse('exam_overview', args=[exam.id]),
    })


def _build_distribution(stats, grade_keys):
    total = stats['total'] or 1
    return [
        {
            'grade': g,
            'range': rng,
            'count': stats['grade_counts'].get(g, 0),
            'pct': round(stats['grade_counts'].get(g, 0) / total * 100),
        }
        for g, rng in grade_keys
    ]


# ── Roster Upload ─────────────────────────────────────────────────────────────

def _pick_col(row, col_lower_map, keys):
    """Match a column using substring containment so that
    'Jina la Mwanafunzi' matches the key 'jina', and 'Student Name'
    matches 'name'.  We walk columns in order and return the first hit,
    which is usually what the user intends.
    """
    # 1) Exact match first (most reliable)
    for k in keys:
        if k in col_lower_map:
            val = str(row[col_lower_map[k]]).strip()
            return '' if val in ('nan', 'None', '') else val
    # 2) Substring match — key must appear as a word-boundary fragment
    #    in the column name (e.g. 'jina' matches 'Jina la Mwanafunzi',
    #    but NOT 'firstname').
    for k in keys:
        for col_lower, col_orig in col_lower_map.items():
            if k in col_lower and k not in ('no', 'id',):  # skip short ambiguous keys
                val = str(row[col_orig]).strip()
                return '' if val in ('nan', 'None', '') else val
    return ''


GENDER_TOKENS = {'m', 'me', 'male', 'kiume', 'f', 'fe', 'female', 'kike'}


def _split_concatenated_names(text):
    """Split concatenated CamelCase names — e.g. 'HalimaAllyMohamedF' → 'Halima Ally Mohamed F'.

    Handles the common case where pdfplumber extracts text without spaces.
    Last single-letter token is treated as gender if it matches M/F.
    """
    import re as _re
    # Insert space before each uppercase letter that follows a lowercase letter
    spaced = _re.sub(r'([a-z])([A-Z])', r'\1 \2', text)
    # Also split digits from letters (e.g. 'John2Smith' → 'John 2 Smith')
    spaced = _re.sub(r'([A-Za-z])(\d)', r'\1 \2', spaced)
    spaced = _re.sub(r'(\d)([A-Za-z])', r'\1 \2', spaced)
    return spaced


def _parse_roster_line(line):
    """Parse one text line: 'FirstName MiddleName LastName Gender' → (first, middle, last, gender)."""
    parts = line.split()
    if len(parts) < 2:
        # May be concatenated (no spaces) — try splitting CamelCase
        if len(line) > 3 and not line[0].isupper():
            return None
        if any(c.isupper() for c in line[1:]):
            line = _split_concatenated_names(line)
            parts = line.split()
    if len(parts) < 2:
        return None

    # Last token is gender if it's a known gender word
    gender = 'M'
    if parts[-1].lower() in GENDER_TOKENS:
        gender = normalize_gender(parts[-1])
        parts = parts[:-1]

    if len(parts) == 1:
        return parts[0].capitalize(), '', 'Unknown', gender
    if len(parts) == 2:
        return parts[0].capitalize(), '', parts[1].capitalize(), gender
    # 3+ parts: first, middle(s), last
    return parts[0].capitalize(), ' '.join(p.capitalize() for p in parts[1:-1]), parts[-1].capitalize(), gender


def _save_student(first, middle, last, gender):
    student, _ = Student.objects.get_or_create(
        first_name=first,
        last_name=last or 'Unknown',
        defaults={'middle_name': middle, 'gender': gender},
    )
    if middle and not student.middle_name:
        student.middle_name = middle
        student.save(update_fields=['middle_name'])
    name_parts = [p for p in [student.first_name, student.middle_name, student.last_name] if p]
    return {'id': student.id, 'name': ' '.join(name_parts)}


def _collect_roster_rows(uploaded_file, is_pdf, form_num=None):
    """Runs the existing PDF/CSV/Excel roster parsing with a no-op
    callback that just records each (first, middle, last, gender) row
    instead of hitting the database per row — the actual save happens
    afterward, once, in bulk (see _bulk_save_students /
    _bulk_save_form_students below). A 1000-row roster used to mean up
    to ~2000 sequential round trips to a remote database via
    get_or_create-per-row; this brings it down to a handful of queries
    total regardless of roster size."""
    collected = []

    def _collector(first, middle, last, gender):
        collected.append((first, middle, last, gender))
        return None

    if is_pdf:
        _parse_pdf_roster(uploaded_file, on_student=_collector)
    else:
        _parse_spreadsheet_roster(uploaded_file, on_student=_collector, form_num=form_num)
    return collected


def _bulk_save_students(parsed_rows):
    """Bulk equivalent of calling _save_student once per row — same
    create-or-get + middle-name-backfill semantics, same first/last-name
    dedup (including within the same file — the first occurrence of a
    name wins, matching get_or_create's per-row behaviour), just as one
    lookup + one bulk_create for the whole roster."""
    if not parsed_rows:
        return []

    rows = [(first, middle, last or 'Unknown', gender) for first, middle, last, gender in parsed_rows]

    first_names = {r[0] for r in rows}
    last_names = {r[2] for r in rows}
    existing = {
        (s.first_name, s.last_name): s
        for s in Student.objects.filter(first_name__in=first_names, last_name__in=last_names)
    }

    new_students = []
    seen = set()
    for first, middle, last, gender in rows:
        key = (first, last)
        if key not in existing and key not in seen:
            seen.add(key)
            new_students.append(Student(first_name=first, middle_name=middle, last_name=last, gender=gender))
    if new_students:
        Student.objects.bulk_create(new_students)
        existing = {
            (s.first_name, s.last_name): s
            for s in Student.objects.filter(first_name__in=first_names, last_name__in=last_names)
        }

    to_update = []
    updated_ids = set()
    for first, middle, last, gender in rows:
        student = existing.get((first, last))
        if student and middle and not student.middle_name and student.id not in updated_ids:
            student.middle_name = middle
            to_update.append(student)
            updated_ids.add(student.id)
    if to_update:
        Student.objects.bulk_update(to_update, ['middle_name'])

    out = []
    for first, middle, last, gender in rows:
        student = existing[(first, last)]
        name_parts = [p for p in [student.first_name, student.middle_name, student.last_name] if p]
        out.append({'id': student.id, 'name': ' '.join(name_parts)})
    return out


def _bulk_save_form_students(school, form_num, parsed_rows):
    """Bulk equivalent of the per-row _save_form_student closure that
    used to live inside upload_form_students — same exact-match dedup
    (first+middle+last, unlike Student which ignores middle) and same
    synthesized admission_no for rows with no free-text one, just batched."""
    import uuid as _uuid

    if not parsed_rows:
        return []

    existing = {
        (fs.first_name, fs.middle_name, fs.last_name): fs
        for fs in FormStudent.objects.filter(school=school, form=form_num)
    }

    new_rows = []
    to_update = []
    seen = set()
    results = []
    for first, middle, last, gender in parsed_rows:
        key = (first, middle, last)
        fs = existing.get(key)
        if fs:
            if fs.gender != gender:
                fs.gender = gender
                to_update.append(fs)
            results.append({'created': False})
        elif key in seen:
            results.append({'created': False})
        else:
            seen.add(key)
            new_rows.append(FormStudent(
                school=school, form=form_num,
                admission_no=f'NA-{_uuid.uuid4().hex[:10]}',
                first_name=first, middle_name=middle, last_name=last, gender=gender,
            ))
            results.append({'created': True})

    if new_rows:
        FormStudent.objects.bulk_create(new_rows)
    if to_update:
        FormStudent.objects.bulk_update(to_update, ['gender'])

    return results


def _parse_pdf_roster(uploaded_file, on_student=_save_student):
    """Extract student rows from a PDF roster using pdfplumber.

    `on_student(first, middle, last, gender)` is called for each row found
    and its return value collected — defaults to _save_student (creates/
    gets a Student, for the per-teacher roster upload), but the Academic's
    form-wide roster upload passes a FormStudent-saving callback instead so
    both uploads share this exact same PDF/CSV parsing logic.
    """
    import pdfplumber, re
    students_out = []
    uploaded_file.seek(0)
    raw_bytes = uploaded_file.read()
    import io
    with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
        for page in pdf.pages:
            # Try table extraction first (structured PDFs)
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        if not row:
                            continue
                        cells = [str(c).strip() for c in row if c and str(c).strip()]
                        if len(cells) < 2:
                            continue
                        # Skip header rows — note: 'no' removed (too aggressive;
                        # matches real names like Noel/Norah).
                        _cell_text = ' '.join(cells).lower()
                        if any(h in _cell_text for h in ('jina la', 'first name', 'last name', 'gender', 'jinsia', 'jinsia ya', '#')):
                            continue
                        # Skip title rows (e.g. "FORM ONE ATTENDANCE LIST SEPT,2026")
                        if any(h in _cell_text for h in ('attendance', 'list', 'sept', 'march', 'may')):
                            continue
                        # Try joining all cells as one line
                        line = ' '.join(cells)
                        parsed = _parse_roster_line(line)
                        if parsed:
                            first, middle, last, gender = parsed
                            if first and first.lower() not in ('nan', 'none', ''):
                                students_out.append(on_student(first, middle, last, gender))
            else:
                # Plain text extraction — each line is one student.
                # Some PDFs contain CSV text (comma-separated) rather than
                # plain text with spaces.  Detect this by checking whether
                # the first non-empty line contains commas.
                text = page.extract_text() or ''
                lines = [l.strip() for l in text.splitlines() if l.strip()]
                if not lines:
                    continue

                # Detect CSV-style: first line has commas and looks like a header
                first_data_line = re.sub(r'^\d+[.)]\s*', '', lines[0]).strip()
                _is_csv = ',' in first_data_line and len(first_data_line.split(',')) >= 2

                # If CSV, detect which column is which from the header
                _col_map = {}  # role → column index
                if _is_csv:
                    header_cells = [c.strip().lower() for c in lines[0].split(',')]
                    for ci, hc in enumerate(header_cells):
                        if hc in ('firstname', 'first name', 'jina la kwanza'):
                            _col_map.setdefault('first', ci)
                        elif hc in ('middlename', 'middle name', 'jina la kati'):
                            _col_map.setdefault('middle', ci)
                        elif hc in ('lastname', 'last name', 'surname', 'jina la mwisho'):
                            _col_map.setdefault('last', ci)
                        elif hc in ('gender', 'jinsia', 'sex'):
                            _col_map.setdefault('gender', ci)
                        elif hc in ('name', 'full name', 'jina', 'student name', 'jina la mwanafunzi'):
                            _col_map.setdefault('full', ci)
                    # If no explicit columns, treat as: name,gender or name...
                    if not _col_map and len(header_cells) >= 2:
                        # Guess: first non-trivial column = name, last = gender
                        for ci, hc in enumerate(header_cells):
                            if hc and hc not in ('namba', 'no', 'no.', '#', 'id'):
                                _col_map.setdefault('full', ci)
                                break
                    _start_row = 1  # skip header
                else:
                    _start_row = 0

                for line in lines[_start_row:]:
                    # Remove leading numbering like "1." "1)" "01."
                    line = re.sub(r'^\d+[.)]\s*', '', line).strip()
                    if not line:
                        continue
                    # Skip obvious header lines (safety net for CSV-detected data too)
                    _line_lower = line.lower()
                    if any(h in _line_lower for h in ('jina la', 'first name', 'last name', 'gender', 'jinsia', 'student')):
                        continue
                    # Skip title lines (e.g. "FORM ONE ATTENDANCE LIST SEPT,2026")
                    if any(h in _line_lower for h in ('attendance', 'list', 'sept', 'march', 'may')):
                        continue

                    if _is_csv:
                        cells = [c.strip() for c in line.split(',')]
                        if len(cells) < 2:
                            continue
                        first = middle = last = ''
                        gender = 'M'
                        if 'first' in _col_map and _col_map['first'] < len(cells):
                            first = cells[_col_map['first']]
                        if 'middle' in _col_map and _col_map['middle'] < len(cells):
                            middle = cells[_col_map['middle']]
                        if 'last' in _col_map and _col_map['last'] < len(cells):
                            last = cells[_col_map['last']]
                        if 'gender' in _col_map and _col_map['gender'] < len(cells):
                            gender = cells[_col_map['gender']]
                        if 'full' in _col_map and _col_map['full'] < len(cells):
                            full_val = cells[_col_map['full']]
                            parts = full_val.split()
                            if not first and len(parts) >= 3:
                                first = parts[0]
                                middle = ' '.join(parts[1:-1])
                                last = parts[-1]
                            elif not first and len(parts) == 2:
                                first = parts[0]
                                last = parts[1]
                            elif not first and parts:
                                first = parts[0]
                        if not first and cells:
                            # Last resort: first non-empty cell
                            for c in cells:
                                if c and c.lower() not in ('nan', 'none', ''):
                                    parsed = _parse_roster_line(c)
                                    if parsed:
                                        first, middle, last, gender = parsed
                                        break
                    else:
                        parsed = _parse_roster_line(line)
                        if not parsed:
                            continue
                        first, middle, last, gender = parsed

                    if first and first.lower() not in ('nan', 'none', ''):
                        students_out.append(on_student(
                            first.strip(), (middle or '').strip(), last.strip() or 'Unknown',
                            normalize_gender(gender)
                        ))
    return students_out

# Header detection keywords — a row containing any of these is
# considered the actual column-header row (not a title row).
_HEADER_KEYWORDS = frozenset({
    's/n', 'sn', '#', 'no', 'no.', 'namba',  # serial number
    'name', 'jina', 'majina', 'mwanafunzi',    # name
    'gender', 'sex', 'jinsia', 'jinsia ya',     # gender
    'signature', 'sahihi',                       # signature
    'score', 'alama', 'mistaari',               # score/marks
    'admission', 'reg no', 'register',           # admission no
})


def _find_header_row(raw_df, max_scan=10):
    """Scan the first *max_scan* rows of a headless DataFrame to find
    the actual header row.  Returns the integer index of that row.

    Many Tanzanian school spreadsheets start with a title row like
    "FORM ONE ATTENDANCE LIST SEPT,2026" before the real header
    (S/N | NAME | ... | SEX).  pandas treats row 0 as column names
    by default, which breaks column detection.
    """
    n_rows = min(len(raw_df), max_scan)
    best_idx = -1
    best_score = 0
    for i in range(n_rows):
        vals = ' '.join(str(v).strip().lower() for v in raw_df.iloc[i] if pd.notna(v))
        if not vals.strip():
            continue
        tokens = set(vals.split())
        hits = sum(1 for kw in _HEADER_KEYWORDS if kw in tokens or kw in vals)
        if hits >= 2 and hits > best_score:  # need at least 2 keyword hits
            best_score = hits
            best_idx = i
    return best_idx if best_idx >= 0 else 0


_GENDER_COL_KEYWORDS = frozenset({'gender', 'sex', 'jinsia', 'jinsia ya'})
_NAME_COL_KEYWORDS = frozenset({'name', 'jina', 'majina', 'mwanafunzi', 'names'})


def _read_best_excel_sheet(uploaded_file, form_num=None):
    """For multi-sheet Excel files, find the sheet that looks most like a
    student roster (has both a name column AND a gender/sex column).

    Many Tanzanian school Excel files contain 10-20+ sheets: marks per
    subject, attendance lists per room, etc.  pandas reads only the first
    sheet by default, which may be a marks sheet with no SEX column —
    causing all students to default to gender='M' and the first name
    to be garbled.

    When *form_num* is provided (1-6), sheets whose title row mentions
    that form (e.g. "FORM THREE ATTENDANCE LIST") get a large score
    bonus so the correct form's roster is selected.

    Returns a cleaned DataFrame ready for column-name detection.
    """
    uploaded_file.seek(0)
    xls = pd.ExcelFile(uploaded_file)
    sheet_names = xls.sheet_names

    # Map form numbers to title keywords for matching
    _FORM_TITLE_MAP = {
        1: ('form one', 'form 1', 'fomu 1', 'fomu ya 1'),
        2: ('form two', 'form 2', 'fomu 2', 'fomu ya 2'),
        3: ('form three', 'form 3', 'fomu 3', 'fomu ya 3'),
        4: ('form four', 'form 4', 'fomu 4', 'fomu ya 4'),
        5: ('form five', 'form 5', 'fomu 5', 'fomu ya 5'),
        6: ('form six', 'form 6', 'fomu 6', 'fomu ya 6'),
    }
    form_keywords = _FORM_TITLE_MAP.get(form_num, ()) if form_num else ()

    best_df = None
    best_score = -1
    best_header_idx = -1

    for sname in sheet_names:
        uploaded_file.seek(0)
        raw = pd.read_excel(xls, sheet_name=sname, header=None)
        if raw.empty or len(raw) < 2:
            continue

        header_idx = _find_header_row(raw)
        if header_idx < 0:
            header_idx = 0

        # Peek at column names from the detected header row
        row_vals = [str(v).strip().lower() for v in raw.iloc[header_idx] if pd.notna(v)]
        all_text = ' '.join(row_vals)

        has_name = any(kw in all_text for kw in _NAME_COL_KEYWORDS)
        has_gender = any(kw in all_text for kw in _GENDER_COL_KEYWORDS)

        # Score: +10 for name, +15 for gender, +2 for each extra header keyword
        score = 0
        if has_name:
            score += 10
        if has_gender:
            score += 15
        score += sum(2 for kw in _HEADER_KEYWORDS if kw in all_text)

        # Big bonus if the title row matches the requested form number
        if form_keywords and header_idx >= 0:
            title_row = ' '.join(str(v).strip().lower() for v in raw.iloc[header_idx - 1] if pd.notna(v)) if header_idx > 0 else ''
            # Also check sheet name itself
            title_text = title_row + ' ' + sname.lower()
            if any(kw in title_text for kw in form_keywords):
                score += 50  # large bonus to prefer this sheet

        # Prefer sheets with both name AND gender columns
        if score > best_score:
            best_score = score
            best_header_idx = header_idx
            # Build a small temp df to return later
            raw.columns = [str(c).strip() for c in raw.iloc[header_idx]]
            best_df = raw.iloc[header_idx + 1:].reset_index(drop=True)

    xls.close()

    if best_df is not None and best_score >= 10:  # at least has a name column
        return best_df

    # Fallback: read the first sheet with default pandas behaviour
    uploaded_file.seek(0)
    return pd.read_excel(uploaded_file)


def _parse_spreadsheet_roster(uploaded_file, on_student=_save_student, form_num=None):
    """Parse CSV or Excel roster with flexible column detection.

    See _parse_pdf_roster for what `on_student` is for.

    Handles Excel/CSV files that have a **title row** before the actual
    header (e.g. "FORM ONE ATTENDANCE LIST SEPT,2026" followed by
    S/N | NAME | SIGNATURE | SCORE | SEX).  pandas normally treats
    the first row as column headers, which breaks column detection.
    """
    import pandas as pd
    uploaded_file.seek(0)
    fname = uploaded_file.name.lower()

    if fname.endswith('.csv'):
        df = pd.read_csv(uploaded_file)
    else:
        df = _read_best_excel_sheet(uploaded_file, form_num=form_num)

    # Clean up column names: strip extra whitespace runs and trailing
    # annotations that some templates add (e.g. "NAME     ROOM A" → "NAME")
    df.columns = [re.sub(r'\s{2,}.*$', '', c).strip() for c in df.columns]
    df.columns = [str(c).strip() for c in df.columns]
    col_lower = {c.lower(): c for c in df.columns}

    students_out = []
    for _, row in df.iterrows():
        first  = _pick_col(row, col_lower, ['first name', 'firstname', 'jina la kwanza'])
        middle = _pick_col(row, col_lower, ['middle name', 'middlename', 'jina la kati', 'jina la pili'])
        last   = _pick_col(row, col_lower, ['last name', 'lastname', 'surname', 'jina la mwisho', 'ukoo'])
        full   = _pick_col(row, col_lower, [
            'name', 'full name', 'jina kamili', 'majina', 'jina',
            'student name', 'jina la mwanafunzi', 'jina kamili la mwanafunzi',
        ])

        if not first and full:
            parts = full.split()
            if len(parts) >= 3:
                first  = parts[0].capitalize()
                middle = ' '.join(p.capitalize() for p in parts[1:-1])
                last   = parts[-1].capitalize()
            elif len(parts) == 2:
                first = parts[0].capitalize()
                last  = parts[1].capitalize()
            elif parts:
                first = parts[0].capitalize()

        # Last-resort fallback: walk columns left-to-right looking for
        # any non-numeric text value that contains at least 2 words.
        # This catches spreadsheets where the name column has a header
        # the detection above doesn't know about (e.g. 'Majina', 'Full',
        # 'Mwanafunzi').
        if not first:
            _SKIP_COLS = {'namba', 'no.', 'no', '#', 'id', 's/n', 'sirina', 'register', 'reg no'}
            for ci in range(len(df.columns)):
                cname = df.columns[ci].strip().lower()
                if cname in _SKIP_COLS:
                    continue
                val = str(row.iloc[ci]).strip()
                if val in ('nan', 'None', '', '0'):
                    continue
                parsed = _parse_roster_line(val)
                if parsed:
                    first, middle, last, _ = parsed
                    break

        if not first or first in ('nan', 'None', ''):
            continue

        gender_raw = _pick_col(row, col_lower, ['gender', 'jinsia', 'sex', 'jinsia ya mwanafunzi']) or 'M'
        students_out.append(on_student(
            first.strip().capitalize(),
            (middle or '').strip().capitalize(),
            (last or 'Unknown').strip().capitalize(),
            normalize_gender(gender_raw),
        ))
    return students_out


@login_required
@require_POST
def upload_roster(request):
    """Accept PDF, CSV, or Excel roster → create/get Student objects → return IDs.

    When exam_id + subject_id are passed, the roster is saved to StoredRoster
    so the teacher can reload it later from the dashboard.
    """
    from .models import StoredRoster
    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return JsonResponse({'error': 'Hakuna faili lililotumwa.'}, status=400)
    try:
        fname = uploaded_file.name.lower()
        rows = _collect_roster_rows(uploaded_file, is_pdf=fname.endswith('.pdf'))
        students_out = _bulk_save_students(rows)

        if not students_out:
            return JsonResponse({
                'error': 'Hakuna wanafunzi waliopatikana. Muundo unaoeleweka: (a) PDF/CSV ya maneno kwa mstari: Halima Ally Mohamed F \n (b) CSV/Excel zenye column: First Name, Middle Name, Last Name, Gender \n (c) CSV/Excel yenye column moja ya jina kamili: Name, Gender'
            }, status=400)

        # Optionally persist roster for later access
        exam_id = request.POST.get('exam_id')
        subject_id = request.POST.get('subject_id')
        if exam_id and subject_id:
            try:
                exam = Exam.objects.filter(id=exam_id).first()
                subject = Subject.objects.filter(id=subject_id).first()
                if exam and subject:
                    StoredRoster.objects.update_or_create(
                        teacher=request.user,
                        exam=exam,
                        subject=subject,
                        defaults={
                            'students': students_out,
                            'student_count': len(students_out),
                            'source_file': uploaded_file.name,
                            'source_format': 'pdf' if fname.endswith('.pdf') else 'csv',
                        },
                    )
            except Exception:
                pass  # roster save is best-effort

        return JsonResponse({'students': students_out, 'count': len(students_out)})
    except Exception as exc:
        return JsonResponse({'error': f'Hitilafu ya faili: {exc}'}, status=400)


@login_required
def download_roster_template(request):
    """Download a ready-to-fill class list template (CSV or PDF)."""
    fmt = request.GET.get('fmt', 'csv').lower()

    # ── CSV template ──
    if fmt == 'csv':
        import csv, io
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(['First Name', 'Middle Name', 'Last Name', 'Gender'])
        w.writerow(['Halima', 'Ally', 'Mohamed', 'F'])
        w.writerow(['Juma', 'Said', 'Hassan', 'M'])
        w.writerow(['Aisha', '', 'Khamis', 'F'])
        w.writerow(['', '', '', ''])  # empty row for teacher to start
        response = HttpResponse(buf.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="Orodha_Ya_Wanafunzi.csv"'
        return response

    # ── PDF template ──
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    import io as _io

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm,
                            leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Title'], fontSize=16, spaceAfter=6,
                                  textColor=colors.HexColor('#1F7A3D'))
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, spaceAfter=10,
                                textColor=colors.HexColor('#333333'))
    hint_style = ParagraphStyle('Hint', parent=styles['Normal'], fontSize=9, spaceAfter=4,
                                 textColor=colors.HexColor('#555555'))
    field_style = ParagraphStyle('Field', parent=styles['Normal'], fontSize=10, spaceAfter=8)

    GREEN = colors.HexColor('#1F7A3D')
    GOLD = colors.HexColor('#D9A441')
    GREY = colors.HexColor('#F2F4F7')

    elements = []
    elements.append(Paragraph('ORODHA YA WANAFUNZI — CLASS LIST', title_style))
    elements.append(Paragraph('Pakua, jaza kwa mkono au kompyuta, kisha piga picha / scan na upakie kwenye mfumo.', sub_style))
    elements.append(Spacer(1, 4*mm))

    # School / class / year fields
    _field = [
        ['Shule / School:', '____________________________', 'Mwaka / Year:', '________'],
        ['Darasa / Class:',  '____________________________', 'Somo / Subject:', '________'],
        ['Mwalimu / Teacher:', '__________________________', '', ''],
    ]
    ft = Table(_field, colWidths=[3.2*cm, 7*cm, 3.2*cm, 3.5*cm])
    ft.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(ft)
    elements.append(Spacer(1, 6*mm))

    # Instructions
    elements.append(Paragraph('<b>Maelekezo / Instructions:</b>', hint_style))
    elements.append(Paragraph("\u2022 Andika jina la kila mwanafunzi kwenye mstari wake / Write each student\u2019s name on their own row.", hint_style))
    elements.append(Paragraph('• Jinsia: F = Mwanamke (Female), M = Mwanaume (Male).', hint_style))
    elements.append(Paragraph('• Jina la Kati (Middle Name) — kama hakuna, acha tupu / leave blank.', hint_style))
    elements.append(Paragraph('• Fomu hii inatumika kwa PDF, Excel, au CSV / Use this form for PDF, Excel, or CSV uploads.', hint_style))
    elements.append(Spacer(1, 6*mm))

    # Student table
    hdr = ['#', 'Jina la Kwanza / First Name', 'Jina la Kati / Middle Name', 'Jina la Mwisho / Last Name', 'Jinsia / Gender']
    rows = [hdr]
    examples = [
        ['1', 'Halima', 'Ally', 'Mohamed', 'F'],
        ['2', 'Juma', 'Said', 'Hassan', 'M'],
        ['3', 'Aisha', '', 'Khamis', 'F'],
    ]
    for ex in examples:
        rows.append(ex)
    # 17 empty rows for teachers to fill
    for i in range(4, 21):
        rows.append([str(i), '', '', '', ''])

    t = Table(rows, colWidths=[1.2*cm, 3.8*cm, 3.8*cm, 3.8*cm, 2*cm])
    t.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), GREEN),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (-1, 1), (-1, -1), 'CENTER'),
        # Example rows — light green background
        ('BACKGROUND', (0, 1), (-1, 3), colors.HexColor('#E8F5E9')),
        # Alternating rows
        *[('BACKGROUND', (0, i), (-1, i), GREY) for i in range(4, 21, 2)],
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('BOX', (0, 0), (-1, -1), 1.2, GREEN),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)

    doc.build(elements)
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Orodha_Ya_Wanafunzi.pdf"'
    return response


# ── Finalize Exam ─────────────────────────────────────────────────────────────

@academic_required
@require_POST
def finalize_exam(request, exam_id):
    exam = _get_exam_or_404(exam_id, request.user)
    subs = exam.subject_submissions.all()
    total = subs.count()
    submitted = subs.filter(
        status__in=[SubjectSubmission.STATUS_SUBMITTED, SubjectSubmission.STATUS_APPROVED]
    ).count()
    if total == 0 or submitted < total:
        messages.error(
            request,
            f"Bado hayajakamilika: masomo {submitted}/{total} yamewasilishwa. "
            "Matokeo ya jumla yanaweza kuzalishwa tu masomo yote yakishawasilishwa."
        )
        return redirect(reverse('exam_overview', args=[exam.id]))
    recompute_processed_results_for_exam(exam)
    return generate_professional_excel_response(exam)


# ── Academic Dashboard ────────────────────────────────────────────────────────

@academic_required
def academic_dashboard(request):
    """Dashboard for academic officer: exams grouped by Form → Stream.

    Shows Form I A, B, C, … then Form II A, B, C, … etc.  Clicking on a
    form level in the sidebar shows all exams / submissions across streams.
    """
    exams = Exam.objects.filter(school=request.user.school).prefetch_related(
        'subject_submissions__subject'
    ).select_related('school').order_by('form', 'stream', '-year', 'name')

    # Nested: { form_num: { stream: [exam_ctx, …], … }, … }
    forms_map = {}       # form_num → { stream_str: [exam_ctx] }
    form_totals = {}     # form_num → { submitted, approved, total }
    exam_count = 0

    for exam in exams:
        exam_count += 1
        subs = list(exam.subject_submissions.all())
        total = len(subs)
        submitted = sum(
            1 for s in subs
            if s.status in (SubjectSubmission.STATUS_SUBMITTED, SubjectSubmission.STATUS_APPROVED)
        )
        approved = sum(1 for s in subs if s.status == SubjectSubmission.STATUS_APPROVED)
        all_submitted = total > 0 and submitted == total
        all_approved = total > 0 and approved == total
        ready_for_approval = all_submitted and not all_approved

        pending_subs = [s for s in subs if s.status == SubjectSubmission.STATUS_PENDING]
        submitted_subs = [s for s in subs if s.status == SubjectSubmission.STATUS_SUBMITTED]
        approved_subs_list = [s for s in subs if s.status == SubjectSubmission.STATUS_APPROVED]
        returned_subs = [s for s in subs if s.status == SubjectSubmission.STATUS_RETURNED]
        pending_names = ', '.join(s.subject.name for s in pending_subs)

        form_key = exam.form
        stream_key = exam.stream or ''
        if form_key not in forms_map:
            forms_map[form_key] = {}
        if stream_key not in forms_map[form_key]:
            forms_map[form_key][stream_key] = []

        exam_ctx = {
            'exam': exam,
            'submissions': subs,
            'pending_subs': pending_subs,
            'submitted_subs': submitted_subs,
            'approved_subs_list': approved_subs_list,
            'returned_subs': returned_subs,
            'pending_names': pending_names,
            'total': total,
            'submitted': submitted,
            'approved': approved,
            'all_submitted': all_submitted,
            'all_approved': all_approved,
            'ready_for_approval': ready_for_approval,
            'progress_pct': round(submitted / total * 100) if total else 0,
            'approval_pct': round(approved / total * 100) if total else 0,
        }
        forms_map[form_key][stream_key].append(exam_ctx)

        # Accumulate form-level totals
        ft = form_totals.setdefault(form_key, {'submitted': 0, 'approved': 0, 'total': 0})
        ft['submitted'] += submitted
        ft['approved'] += approved
        ft['total'] += total

    # Build sorted structure for template: [(form_num, { stream: [exams] })]
    FORM_LABELS = Exam.FORM_LABELS
    forms_list = []
    for form_num in sorted(forms_map.keys()):
        streams = forms_map[form_num]
        form_label = FORM_LABELS.get(form_num, f'Form {form_num}')
        ft = form_totals.get(form_num, {'submitted': 0, 'approved': 0, 'total': 0})
        forms_list.append({
            'form_num': form_num,
            'form_label': form_label,
            'streams': streams,
            'total_submitted': ft['submitted'],
            'total_approved': ft['approved'],
            'total_all': ft['total'],
        })

    return render(request, 'results/academic_dashboard.html', {
        'forms_list': forms_list,
        'total_exams': exam_count,
    })


# ── Approve Subject Submission ────────────────────────────────────────────────

@academic_required
@require_POST
def approve_subject(request, exam_id, subject_id):
    exam = _get_exam_or_404(exam_id, request.user)
    subject = get_object_or_404(Subject, id=subject_id)
    sub = get_object_or_404(SubjectSubmission, exam=exam, subject=subject)

    if sub.status != SubjectSubmission.STATUS_SUBMITTED:
        messages.error(request, f"{subject.name} haijawa na hali ya Submitted.")
        return redirect(reverse('academic_dashboard'))

    approved_by = request.user.full_name or request.user.email
    notes = request.POST.get('notes', '').strip()

    sub.status = SubjectSubmission.STATUS_APPROVED
    sub.approved_by = approved_by
    sub.approved_by_user = request.user
    sub.approved_at = timezone.now()
    sub.approval_notes = notes
    sub.save(update_fields=['status', 'approved_by', 'approved_by_user', 'approved_at', 'approval_notes'])
    recompute_processed_results_for_exam(exam)

    messages.success(request, f"Somo la {subject.name} limeidhinishwa.")
    return redirect(reverse('exam_overview', args=[exam.id]))


# ── Return Submission to Teacher ─────────────────────────────────────────────

@academic_required
@require_POST
def return_submission(request, exam_id, subject_id):
    """Mtaaluma anarudisha submission kwa mwalimu kwa ukarabati.

    Mtaaluma anaandika comment (return_notes) inayoonekana na mwalimu.
    Submission hubadilisha status kuwa RETURNED — mwalimu atarekebisha
    na kuwasilisha upya, na submission ya zamani itatolewa.
    """
    exam = _get_exam_or_404(exam_id, request.user)
    subject = get_object_or_404(Subject, id=subject_id)
    sub = get_object_or_404(SubjectSubmission, exam=exam, subject=subject)

    if sub.status not in (SubjectSubmission.STATUS_SUBMITTED,):
        messages.error(request, f"{subject.name} haiwezi kurudishwa (hali: {sub.get_status_display()}).")
        return redirect(reverse('academic_dashboard'))

    notes = request.POST.get('return_notes', '').strip()
    if not notes:
        messages.error(request, "Andika sababu ya kurudisha submission hii.")
        return redirect(reverse('exam_overview', args=[exam.id]))

    sub.status = SubjectSubmission.STATUS_RETURNED
    sub.return_notes = notes
    sub.returned_at = timezone.now()
    sub.returned_by = request.user.full_name or request.user.email
    sub.returned_by_user = request.user
    sub.save(update_fields=[
        'status', 'return_notes', 'returned_at', 'returned_by', 'returned_by_user',
    ])

    messages.warning(
        request,
        f"Somo la {subject.name} limerudishwa kwa mwalimu ({sub.submitted_by}) kwa ukarabati."
    )
    return redirect(reverse('exam_overview', args=[exam.id]))


# ── Bulk Approve All Subjects for an Exam ────────────────────────────────────

@academic_required
@require_POST
def approve_exam_submissions(request, exam_id):
    exam = _get_exam_or_404(exam_id, request.user)
    approved_by = request.user.full_name or request.user.email
    notes = request.POST.get('notes', '').strip()
    now = timezone.now()

    updated = SubjectSubmission.objects.filter(
        exam=exam, status=SubjectSubmission.STATUS_SUBMITTED
    ).update(
        status=SubjectSubmission.STATUS_APPROVED,
        approved_by=approved_by,
        approved_by_user=request.user,
        approved_at=now,
        approval_notes=notes,
    )

    if updated == 0:
        messages.warning(request, "Hakuna masomo yaliyokuwa tayari kwa idhini.")
    else:
        recompute_processed_results_for_exam(exam)
        messages.success(request, f"Masomo {updated} yameidhinishwa. Matokeo yamehesabiwa upya.")

    return redirect(reverse('exam_overview', args=[exam.id]))


# ── Recompute Results (for exams already fully approved) ────────────────────
# approve_subject/approve_exam_submissions only recompute as a side effect of
# a status change to APPROVED — once every subject is already approved there
# was no way to force a fresh recompute (e.g. after a grading-rule fix), so
# AGGT/points/division on an already-processed exam stayed stale forever.

@academic_required
@require_POST
def recompute_exam_results(request, exam_id):
    exam = _get_exam_or_404(exam_id, request.user)
    recompute_processed_results_for_exam(exam)
    messages.success(request, "Matokeo yamehesabiwa upya.")
    return redirect(reverse('exam_overview', args=[exam.id]))


# ── Form-Level Results View ───────────────────────────────────────────────────

@login_required
def form_results(request, form_num):
    """Results for all approved exams of a given form level."""
    exams = list(Exam.objects.filter(form=form_num, school=request.user.school).order_by('-year', 'name'))
    is_academic = getattr(request.user, 'is_academic', False)

    from .services.subject_pdf_service import get_grade_keys_for_form

    if not exams:
        return render(request, 'results/form_results.html', {
            'form_num': form_num,
            'exams_ctx': [],
            'form_label': f'Form {form_num}' if form_num <= 4 else f'Form {form_num} (Advanced)',
            'is_academic': is_academic,
        })

    exam_ids = [e.id for e in exams]

    # ── Batch-load all submission counts in one query ──
    sub_counts = SubjectSubmission.objects.filter(
        exam_id__in=exam_ids
    ).values('exam_id').annotate(
        total_subs=Count('id'),
        approved_subs=Count(Case(
            When(status=SubjectSubmission.STATUS_APPROVED, then=1),
            output_field=IntegerField(),
        )),
        submitted_subs=Count(Case(
            When(status__in=[SubjectSubmission.STATUS_SUBMITTED, SubjectSubmission.STATUS_APPROVED], then=1),
            output_field=IntegerField(),
        )),
    )
    sub_counts_map = {r['exam_id']: r for r in sub_counts}

    # ── Batch-load subjects, processed results, and exam results once ──
    all_subjects = Subject.objects.filter(
        examresult__exam_id__in=exam_ids
    ).distinct().order_by('name')
    all_subjects_list = list(all_subjects)
    all_subject_ids = [s.id for s in all_subjects_list]

    processed_qs = ProcessedResult.objects.filter(
        exam_id__in=exam_ids
    ).select_related('student').order_by('position')
    processed_by_exam = {}
    for pr in processed_qs:
        processed_by_exam.setdefault(pr.exam_id, []).append(pr)

    # Exam results — build score_lookup keyed by (student_id, subject_id) -> score,
    # plus per-exam grouping for grade_lookup
    score_lookup_global = {}
    for er in ExamResult.objects.filter(
        exam_id__in=exam_ids, subject_id__in=all_subject_ids
    ).only('student_id', 'subject_id', 'score', 'exam_id'):
        score_lookup_global[(er.exam_id, er.student_id, er.subject_id)] = er.score

    # ── Batch-load PS submissions once ──
    from .models import PrintSubmission
    ps_map = {}
    for ps in PrintSubmission.objects.filter(
        exam_id__in=exam_ids, school=request.user.school
    ).order_by('exam_id', '-submitted_at'):
        ps_map.setdefault(ps.exam_id, ps)

    # ── Build per-exam context from preloaded data ──
    from collections import defaultdict
    subjects_by_exam = defaultdict(list)
    for (eid, sid, _), _ in score_lookup_global.items():
        pass  # just ensuring it's iterated
    # subjects per exam: which subjects have results
    _subjects_by_exam = defaultdict(set)
    for (eid, sid, _), _ in score_lookup_global.items():
        _subjects_by_exam[eid].add(sid)
    subjects_by_id = {s.id: s for s in all_subjects_list}

    exams_ctx = []
    for exam in exams:
        counts = sub_counts_map.get(exam.id, {'total_subs': 0, 'approved_subs': 0, 'submitted_subs': 0})
        total_subs = counts['total_subs']
        approved_subs = counts['approved_subs']
        submitted_subs = counts['submitted_subs']
        all_submitted = total_subs > 0 and submitted_subs == total_subs
        all_approved = total_subs > 0 and approved_subs == total_subs

        exam_subject_ids = _subjects_by_exam.get(exam.id, set())
        exam_subjects = [subjects_by_id[sid] for sid in sorted(exam_subject_ids) if sid in subjects_by_id]
        exam_processed = processed_by_exam.get(exam.id, [])

        # Build score_lookup and grade_lookup for this exam
        score_lookup = {}
        grade_lookup = {}
        for (eid, sid, subj_id), score in score_lookup_global.items():
            if eid == exam.id:
                score_lookup[(sid, subj_id)] = score
                grade_lookup.setdefault(sid, {})[subj_id] = get_grade_for_form(score, exam.form)
        grade_key = get_grade_keys_for_form(exam.form)

        exams_ctx.append({
            'exam': exam,
            'total_subs': total_subs,
            'approved_subs': approved_subs,
            'submitted_subs': submitted_subs,
            'all_submitted': all_submitted,
            'all_approved': all_approved,
            'processed_results': exam_processed,
            'subjects': exam_subjects,
            'grade_lookup': grade_lookup,
            'grade_key': grade_key,
            'pdf_url': reverse('generate_results_pdf', args=[exam.id]),
            'bulk_pdf_url': reverse('generate_bulk_student_results_pdf', args=[exam.id]),
            'overview_url': reverse('exam_overview', args=[exam.id]),
            'approve_all_url': reverse('approve_exam_submissions', args=[exam.id]) if is_academic else None,
            'recompute_url': reverse('recompute_exam_results', args=[exam.id]) if is_academic else None,
            'submit_ps_url': reverse('submit_exam_to_ps', args=[exam.id]) if is_academic else None,
            'ps_submission': ps_map.get(exam.id),
        })

    return render(request, 'results/form_results.html', {
        'form_num': form_num,
        'exams_ctx': exams_ctx,
        'form_label': f'Form {form_num}' if form_num <= 4 else f'Form {form_num} (Advanced)',
        'is_academic': is_academic,
    })


# ── Form-Level Excel Export ───────────────────────────────────────────────────

@academic_required
def form_results_excel(request, form_num):
    """Single Excel workbook with one sheet per exam for the given form."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY, GOLD, WHITE, GREY = "FF1F7A3D", "FFD9A441", "FFFFFFFF", "FFF2F4F7"

    def _fill(hex_c):
        return PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")

    def _border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def _score_color(score):
        if score is None:
            return None, None
        if score >= 75: return "FFC6F4D6", "FF145A32"   # A
        if score >= 65: return "FFD5F5E3", "FF1E8449"   # B
        if score >= 45: return "FFFFF9C4", "FF7D6608"   # C
        if score >= 30: return "FFFDEBD0", "FF784212"   # D
        return "FFFADBD8", "FF922B21"                   # F

    exams = list(Exam.objects.filter(form=form_num, school=request.user.school).order_by('-year', 'name'))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── Batch-load all data upfront (same approach as form_results) ──
    exam_ids = [e.id for e in exams]
    all_subjects_list = list(Subject.objects.filter(
        examresult__exam_id__in=exam_ids
    ).distinct().order_by('name'))
    all_subject_ids = [s.id for s in all_subjects_list]
    subjects_by_id = {s.id: s for s in all_subjects_list}

    processed_by_exam = {}
    for pr in ProcessedResult.objects.filter(
        exam_id__in=exam_ids
    ).select_related('student').order_by('position'):
        processed_by_exam.setdefault(pr.exam_id, []).append(pr)

    score_lookup_global = {}
    for er in ExamResult.objects.filter(
        exam_id__in=exam_ids, subject_id__in=all_subject_ids
    ).only('exam_id', 'student_id', 'subject_id', 'score'):
        score_lookup_global[(er.exam_id, er.student_id, er.subject_id)] = er.score

    _subjects_by_exam = defaultdict(set)
    for (eid, sid, _) in score_lookup_global:
        _subjects_by_exam[eid].add(sid)

    for exam in exams:
        subjects = [subjects_by_id[sid] for sid in sorted(_subjects_by_exam.get(exam.id, set())) if sid in subjects_by_id]
        results = processed_by_exam.get(exam.id, [])
        score_lookup = {(sid, subj_id): score for (eid, sid, subj_id), score in score_lookup_global.items() if eid == exam.id}

        sheet_title = f"{exam.name[:25]} {exam.year}"[:31]
        ws = wb.create_sheet(title=sheet_title)

        total_cols = 3 + len(subjects) + 4  # POS JINA JINSIA + subjects + JUMLA WASTANI DARAJA POINTI

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        c = ws.cell(row=1, column=1)
        school = exam.school_name or 'SHULE'
        c.value = f"{school.upper()} — {exam.get_exam_type_display().upper()} {exam.year} — FORM {exam.form}"
        c.font = Font(bold=True, size=13, color=WHITE, name='Calibri')
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 26

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        c2 = ws.cell(row=2, column=1)
        c2.value = exam.name
        c2.font = Font(bold=False, size=10, color=GOLD, name='Calibri')
        c2.fill = _fill(NAVY)
        c2.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 16

        # Header row
        headers = ['POS', 'JINA', 'JINSIA'] + [s.name.upper() for s in subjects] + ['JUMLA', 'WASTANI', 'DARAJA', 'POINTI']
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = Font(color=WHITE, bold=True, name='Calibri', size=9)
            cell.fill = _fill(NAVY)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = _border()
        ws.row_dimensions[3].height = 20
        ws.freeze_panes = ws.cell(row=4, column=1)

        # Data rows
        for ri, result in enumerate(results):
            row = 4 + ri
            st = result.student
            name = ' '.join(p for p in [st.first_name, st.middle_name or '', st.last_name] if p)
            row_fill = _fill(GREY) if ri % 2 == 0 else None

            ws.cell(row=row, column=1, value=result.position)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=st.gender)

            for si, subj in enumerate(subjects):
                score = score_lookup.get((st.id, subj.id))
                col = 4 + si
                cell = ws.cell(row=row, column=col)
                if score is not None:
                    cell.value = score
                    bg, fg = _score_color(score)
                    if bg: cell.fill = _fill(bg)
                    if fg: cell.font = Font(bold=True, name='Calibri', size=9, color=fg)
                else:
                    cell.value = '—'

            base = 4 + len(subjects)
            ws.cell(row=row, column=base).value = result.total_score
            ws.cell(row=row, column=base).font = Font(bold=True, name='Calibri', size=9, color=NAVY)
            avg_cell = ws.cell(row=row, column=base + 1, value=float(result.average_score))
            avg_cell.number_format = '0.00'
            ws.cell(row=row, column=base + 2, value=result.division)
            ws.cell(row=row, column=base + 3, value=result.points)

            for ci in range(1, total_cols + 1):
                cell = ws.cell(row=row, column=ci)
                cell.border = _border()
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if row_fill and not cell.fill.fgColor.rgb not in ('00000000', '000000'):
                    if not cell.fill or cell.fill.fill_type == 'none':
                        cell.fill = row_fill
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center')

        # Summary row
        if results:
            last = 4 + len(results)
            ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=3)
            sc = ws.cell(row=last, column=1, value=f"Jumla ya Wanafunzi: {len(results)}")
            sc.font = Font(bold=True, name='Calibri', size=9, color=WHITE)
            sc.fill = _fill(NAVY)
            sc.alignment = Alignment(horizontal='left', vertical='center')

        # Auto column widths
        for col in ws.columns:
            max_len = 8
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)
        ws.column_dimensions['B'].width = 28

    if not wb.sheetnames:
        ws = wb.create_sheet("Hakuna Data")
        ws.cell(row=1, column=1, value=f"Hakuna mitihani ya Form {form_num} bado.")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    label = f'Form_{form_num}'
    response['Content-Disposition'] = f'attachment; filename="Matokeo_{label}.xlsx"'
    wb.save(response)
    return response


# ── My School ─────────────────────────────────────────────────────────────────
# Schools themselves are registered by the site's system administrator via
# Django admin (each with its first Academic account). An academic officer
# only ever sees and manages their own school — they can no longer browse or
# create other schools from here.

@academic_required
def school_setup(request):
    """Read-only info about the academic officer's own school."""
    school = request.user.school
    if not school:
        messages.error(
            request,
            "Akaunti yako haijapangiwa shule. Wasiliana na msimamizi wa mfumo (system admin)."
        )
    return render(request, 'results/school_setup.html', {'school': school})


# ── School Subjects Management ────────────────────────────────────────────────

@academic_required
def school_subjects(request):
    """Add/remove subjects taught at the academic officer's own school."""
    school = request.user.school
    if not school:
        messages.error(request, "Akaunti yako haijapangiwa shule. Wasiliana na msimamizi wa mfumo.")
        return redirect(reverse('home'))

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add':
            subject_name = request.POST.get('subject_name', '').strip()
            form_levels = request.POST.get('form_levels', '1,2,3,4').strip()
            if subject_name:
                subject = safe_get_or_create_subject(subject_name)
                ss, created = SchoolSubject.objects.get_or_create(
                    school=school,
                    subject=subject,
                    defaults={'form_levels': form_levels},
                )
                if created:
                    messages.success(request, f"Somo '{subject.name}' limeongezwa.")
                else:
                    messages.info(request, f"Somo '{subject.name}' tayari lipo.")
            else:
                messages.error(request, "Andika jina la somo.")

        elif action == 'remove':
            ss_id = request.POST.get('school_subject_id')
            if ss_id:
                SchoolSubject.objects.filter(id=ss_id, school=school).delete()
                messages.success(request, "Somo limeondolewa.")

        return redirect(reverse('school_subjects'))

    school_subjects_qs = school.school_subjects.select_related('subject').order_by('subject__name')
    existing_ids = school_subjects_qs.values_list('subject_id', flat=True)
    available_subjects = Subject.objects.exclude(id__in=existing_ids).order_by('name')

    return render(request, 'results/school_subjects.html', {
        'school': school,
        'school_subjects': school_subjects_qs,
        'available_subjects': available_subjects,
        'common_subjects': COMMON_SUBJECTS,
    })


# ── Create Exam for School ────────────────────────────────────────────────────

@academic_required
def create_exam_for_school(request):
    """Create an Exam for the academic officer's own school and auto-create
    PENDING SubjectSubmissions for every subject the school teaches."""
    school = request.user.school
    if not school:
        messages.error(request, "Akaunti yako haijapangiwa shule. Wasiliana na msimamizi wa mfumo.")
        return redirect(reverse('home'))

    if request.method == 'POST':
        name = request.POST.get('exam_name', '').strip()
        year = int(request.POST.get('exam_year', timezone.now().year))
        form_level = int(request.POST.get('exam_form', 4))
        stream = request.POST.get('exam_stream', '').strip()
        exam_type = request.POST.get('exam_type', 'TERMINAL')

        if not name:
            messages.error(request, "Tafadhali weka jina la mtihani.")
            return redirect(request.path)

        exam, created = Exam.objects.get_or_create(
            name=name,
            year=year,
            form=form_level,
            stream=stream,
            exam_type=exam_type,
            school=school,
            defaults={'school_name': school.name},
        )
        if not created:
            messages.info(request, f"Mtihani '{exam.name}' tayari upo.")
        else:
            # Auto-create PENDING SubjectSubmissions for each school subject
            school_subjs = school.school_subjects.select_related('subject').all()
            for ss in school_subjs:
                SubjectSubmission.objects.get_or_create(
                    exam=exam,
                    subject=ss.subject,
                    defaults={'status': SubjectSubmission.STATUS_PENDING},
                )
            messages.success(
                request,
                f"Mtihani '{exam.name}' umeundwa. Masomo {school_subjs.count()} yanangojea kupakiwa."
            )

        return redirect(reverse('exam_overview', args=[exam.id]))

    school_subjects_qs = school.school_subjects.select_related('subject').order_by('subject__name')
    current_year = timezone.now().year

    return render(request, 'results/create_exam.html', {
        'school': school,
        'school_subjects': school_subjects_qs,
        'exam_type_choices': _EXAM_TYPE_CHOICES,
        'current_year': current_year,
        'year_range': range(current_year - 2, current_year + 3),
    })


# ── Delete Exam ──────────────────────────────────────────────────────────────

@academic_required
@require_POST
def delete_exam(request, exam_id):
    """Delete an exam and all its related data (results, submissions, etc.).

    Only the Academic Officer for the same school can delete. This is
    irreversible — we show a confirmation modal on the frontend.
    """
    exam = _get_exam_or_404(exam_id, request.user)
    exam_name = exam.name
    exam.delete()  # CASCADE deletes ExamResult, SubjectSubmission, ProcessedResult, etc.
    messages.success(request, f"Mtihani '{exam_name}' umefutwa.")
    return redirect(reverse('academic_dashboard'))


# ── Teacher Dashboard ─────────────────────────────────────────────────────────

@teacher_required
def teacher_dashboard(request):
    """Walimu wanaona masomo yao pekee ya kupakia (submission status kwa kila exam).
    Pia onyesha orodha zilizohifadhiwa (stored rosters) kwa urahisi wa kujaza alama."""
    from .models import StoredRoster
    teacher_subject_ids = set(request.user.subjects.values_list('id', flat=True))

    exams = Exam.objects.filter(school=request.user.school).prefetch_related(
        'subject_submissions__subject'
    ).order_by('-year', 'form', 'name')

    roster_lookup = {
        (r.exam_id, r.subject_id): r
        for r in StoredRoster.objects.filter(teacher=request.user).only(
            'exam_id', 'subject_id', 'student_count'
        )
    }

    exams_ctx = []
    for exam in exams:
        subs = [s for s in exam.subject_submissions.all() if s.subject_id in teacher_subject_ids]
        if not subs:
            continue
        pending = [s for s in subs if s.status == SubjectSubmission.STATUS_PENDING]
        submitted = [s for s in subs if s.status == SubjectSubmission.STATUS_SUBMITTED]
        approved = [s for s in subs if s.status == SubjectSubmission.STATUS_APPROVED]
        returned = [s for s in subs if s.status == SubjectSubmission.STATUS_RETURNED]
        for s in submitted + approved:
            s.pdf_url = reverse('subject_pdf', args=[exam.id, s.subject_id])
            s.summary_url = reverse('subject_summary', args=[exam.id, s.subject_id])
        # A teacher can still edit their own marks after submitting — even
        # once approved, in which case re-saving/re-submitting flips the
        # status back to SUBMITTED so the academic officer knows to look
        # at it again. Every status below PENDING gets a way back in.
        for s in submitted + approved + returned:
            s.marks_url = reverse('marks_entry') + f'?exam={exam.id}&subject={s.subject_id}'
        # Attach stored roster info for pending subjects
        for s in pending:
            roster = roster_lookup.get((exam.id, s.subject_id))
            s.has_roster = roster is not None
            s.roster_count = roster.student_count if roster else 0
            s.marks_url = reverse('marks_entry') + f'?exam={exam.id}&subject={s.subject_id}'
        exams_ctx.append({
            'exam': exam,
            'pending': pending,
            'submitted': submitted,
            'approved': approved,
            'returned': returned,
            'total': len(subs),
        })

    # Stored rosters summary
    stored_rosters = StoredRoster.objects.filter(
        teacher=request.user
    ).select_related('exam', 'subject').order_by('-created_at')[:20]

    return render(request, 'results/teacher_dashboard.html', {
        'exams_ctx': exams_ctx,
        'has_subjects': bool(teacher_subject_ids),
        'stored_rosters': stored_rosters,
    })


# ── Self-Service Subject Selection — teacher picks their own subject(s) ──────

@login_required
def select_my_subjects(request):
    """Pick the subject(s) you teach yourself — no one else needs to assign
    them. Open to both roles: academic officers often teach a subject too,
    on top of their admin duties."""
    if request.method == 'POST':
        form = TeacherSelfSubjectsForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Masomo yako yamesasishwa.")
            return redirect('academic_dashboard' if request.user.is_academic else 'teacher_dashboard')
    else:
        form = TeacherSelfSubjectsForm(instance=request.user)

    return render(request, 'results/select_subjects.html', {'form': form})


# ── Personal (Binafsi) Upload — private scratch tool, not part of the official exam ──

@teacher_required
def personal_upload(request):
    """Mwalimu anapakia alama za somo lake mwenyewe na kupata PDF papo hapo.

    Hii haihitaji Mtihani (Exam) uliotengenezwa na Afisa Taaluma, na
    haiathiri matokeo ya jumla ya shule kwa vyovyote vile.
    """
    teacher_subjects = request.user.subjects.all().order_by('name')
    recent_uploads = PersonalUpload.objects.filter(teacher=request.user).select_related('subject')[:10]

    if request.method == 'POST':
        subject_id = request.POST.get('subject_id')
        # Jina la tathmini linachaguliwa kwenye dropdown (aina za mtihani)
        # badala ya kuandikwa — default ni 'OTHER' ikiwa hakuna kilichochaguliwa.
        title_choice = request.POST.get('title', '').strip()
        title_label = dict(Exam.EXAM_TYPE_CHOICES).get(title_choice, '')
        title = title_label or title_choice or 'Matokeo Binafsi'
        uploaded_file = request.FILES.get('file')

        subject = teacher_subjects.filter(pk=subject_id).first()
        if not subject:
            messages.error(request, "Chagua somo unalofundisha.")
            return redirect('personal_upload')
        if not uploaded_file:
            messages.error(request, "Hakuna faili lililochaguliwa.")
            return redirect('personal_upload')

        try:
            rows = parse_name_score_sheet(uploaded_file)
        except ValidationError as e:
            messages.error(request, str(e))
            return redirect('personal_upload')
        except Exception as e:
            messages.error(request, f"Hitilafu ya faili: {e}")
            return redirect('personal_upload')

        if not rows:
            messages.error(request, "Hakuna wanafunzi waliopatikana kwenye faili.")
            return redirect('personal_upload')

        upload = PersonalUpload.objects.create(teacher=request.user, subject=subject, title=title)
        PersonalUploadResult.objects.bulk_create([
            PersonalUploadResult(upload=upload, student_name=name, score=score) for name, score in rows
        ])
        messages.success(request, f"Alama {len(rows)} zimepakiwa. Pakua PDF ya matokeo hapa chini.")
        return redirect('personal_upload')

    return render(request, 'results/personal_upload.html', {
        'teacher_subjects': teacher_subjects,
        'recent_uploads': recent_uploads,
        'exam_type_choices': Exam.EXAM_TYPE_CHOICES,
    })


@teacher_required
def personal_upload_pdf(request, upload_id):
    upload = get_object_or_404(PersonalUpload, id=upload_id, teacher=request.user)
    lang = request.session.get('ui_lang', 'en')
    return generate_personal_pdf_response(upload, lang=lang)


@teacher_required
def personal_upload_summary(request, upload_id):
    upload = get_object_or_404(PersonalUpload, id=upload_id, teacher=request.user)
    results = list(upload.results.order_by('-score', 'student_name'))
    rows_data = []
    for pos, result in enumerate(results, 1):
        rows_data.append({
            'position': pos,
            'name': result.student_name,
            'score': result.score,
            'grade': get_grade(result.score),
            'gender': None,
        })

    lang = request.session.get('ui_lang', 'en')
    stats = compute_subject_stats(rows_data)
    recommendations = generate_recommendations(stats, subject_name=upload.subject.name, lang=lang)
    uploader_name = upload.teacher.full_name or upload.teacher.email
    subject_word, teacher_word = ('Somo', 'Mwalimu') if lang == 'sw' else ('Subject', 'Teacher')
    distribution = _build_distribution(stats, GRADE_KEYS_OLEVEL)

    return render(request, 'results/results_summary.html', {
        'heading': upload.title,
        'meta_parts': [f"{subject_word}: {upload.subject.name}", f"{teacher_word}: {uploader_name}", upload.created_at.strftime('%d %b %Y')],
        'rows_data': rows_data,
        'stats': stats,
        'recommendations': recommendations,
        'distribution': distribution,
        'pdf_url': reverse('personal_upload_pdf', args=[upload.id]),
        'back_url': reverse('personal_upload'),
    })


# ── Logo Upload ──────────────────────────────────────────────────────────────

@academic_required
def upload_logos(request):
    """Upload school and district logos for the PDF report header."""
    school = request.user.school
    if not school:
        messages.error(request, "Hakuna shule iliyowekwa. Weka shule kwanza.")
        return redirect('home')

    if request.method == 'POST':
        logo_type = request.POST.get('logo_type', '')
        uploaded_file = request.FILES.get('logo_file')

        if not uploaded_file:
            messages.error(request, "Hakuna faili lililochaguliwa.")
            return redirect('upload_logos')

        # Validate file type — SVG is deliberately NOT accepted: the PDF
        # renderer (ReportLab's ImageReader, backed by Pillow) cannot
        # decode SVG at all. It used to be listed as "supported" here but
        # every SVG upload silently failed to ever appear in the PDF, with
        # no error anywhere, because the drawing code swallows the
        # resulting exception. PNG/JPG only, so what's accepted here is
        # guaranteed to actually be drawable.
        allowed_types = ['image/png', 'image/jpeg', 'image/jpg']
        if uploaded_file.content_type not in allowed_types:
            messages.error(request, "Aina ya faili haitambuliki. Tumia PNG au JPG (SVG haiwezi kuonyeshwa kwenye PDF).")
            return redirect('upload_logos')

        # Validate file size (max 2MB)
        if uploaded_file.size > 2 * 1024 * 1024:
            messages.error(request, "Faili ni kubwa sana. Kiwango ni 2MB.")
            return redirect('upload_logos')

        # Verify the bytes are actually a decodable image — a spoofed or
        # corrupt content_type would otherwise pass the check above and
        # then silently fail to render in the PDF with no error at all.
        import io as _io
        from PIL import Image as _PILImage
        file_data = uploaded_file.read()
        try:
            _PILImage.open(_io.BytesIO(file_data)).verify()
        except Exception:
            messages.error(request, "Faili hii si picha inayosomeka. Jaribu picha nyingine ya PNG/JPG.")
            return redirect('upload_logos')

        # Save to School model — both ImageField AND base64 in DB
        # (base64 persists on Railway where filesystem is ephemeral)
        import base64 as _b64
        ext = Path(uploaded_file.name).suffix.lower()
        mime_map = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.gif': 'image/gif'}
        mime = mime_map.get(ext, 'image/png')
        b64_str = f'data:{mime};base64,{_b64.b64encode(file_data).decode("ascii")}'
        uploaded_file.seek(0)  # reset for ImageField save

        if logo_type == 'school':
            school.school_logo = uploaded_file
            school.school_logo_b64 = b64_str
            school.save(update_fields=['school_logo', 'school_logo_b64'])
            logo_name = "ya Shule"
        elif logo_type == 'district':
            school.district_logo = uploaded_file
            school.district_logo_b64 = b64_str
            school.save(update_fields=['district_logo', 'district_logo_b64'])
            logo_name = "ya Halmashauri"
        elif logo_type == 'coa':
            school.coat_of_arms = uploaded_file
            school.coat_of_arms_b64 = b64_str
            school.save(update_fields=['coat_of_arms', 'coat_of_arms_b64'])
            logo_name = "ya Coat of Arms"
        else:
            messages.error(request, "Aina ya logo haijulikani.")
            return redirect('upload_logos')

        messages.success(request, f"Logo {logo_name} imehifadhiwa! Itaonekana kwenye PDF ya matokeo.")
        return redirect('upload_logos')

    return render(request, 'results/upload_logos.html', {
        'school_logo_exists': bool(school.school_logo_b64 or school.school_logo) if school else False,
        'district_logo_exists': bool(school.district_logo_b64 or school.district_logo) if school else False,
        'coa_exists': bool(school.coat_of_arms_b64 or school.coat_of_arms) if school else False,
        'school_logo_url': school.school_logo.url if school and school.school_logo else '',
        'district_logo_url': school.district_logo.url if school and school.district_logo else '',
        'coa_url': school.coat_of_arms.url if school and school.coat_of_arms else '',
    })


# ══════════════════════════════════════════════════════════════════════════════
# FORM STUDENT LIST — Upload + Teacher Dashboard
# ══════════════════════════════════════════════════════════════════════════════

@academic_required
def upload_form_students(request):
    """Academic Officer uploads orodha ya wanafunzi kwa form fulani.

    Same file formats as the per-teacher roster upload (upload_roster):
    PDF (one full name + gender per line, e.g. "Halima Ally Mohamed F"),
    or CSV/Excel with columns First Name | Middle Name | Last Name |
    Gender — parsed by the exact same _parse_pdf_roster/
    _parse_spreadsheet_roster used there, just pointed at a FormStudent-
    saving callback instead of Student, so both uploads accept identically
    formatted files.
    """
    school = request.user.school
    if not school:
        messages.error(request, "Hakuna shule iliyowekwa.")
        return redirect('home')

    form_num = request.GET.get('form') or request.POST.get('form') or ''
    selected_form = int(form_num) if form_num.isdigit() and int(form_num) in (1,2,3,4,5,6) else None

    if request.method == 'POST' and selected_form:
        uploaded_file = request.FILES.get('student_file')
        if not uploaded_file:
            messages.error(request, "Chagua faili la orodha ya wanafunzi.")
            return redirect(f'{reverse("upload_form_students")}?form={selected_form}')

        ext = Path(uploaded_file.name).suffix.lower()

        try:
            if ext == '.pdf':
                rows = _collect_roster_rows(uploaded_file, is_pdf=True, form_num=selected_form)
            elif ext in ('.csv', '.xlsx', '.xls'):
                rows = _collect_roster_rows(uploaded_file, is_pdf=False, form_num=selected_form)
            else:
                messages.error(request, f"Aina ya faili '{ext}' haijulikani. Tumia PDF, CSV au Excel.")
                return redirect(f'{reverse("upload_form_students")}?form={selected_form}')

            saved = _bulk_save_form_students(school, selected_form, rows)
            count = len(saved)
            if count:
                messages.success(request, f"Wanafunzi {count} wa Form {selected_form} wamehifadhiwa!")
            else:
                messages.warning(request, "Hakuna wanafunzi waliohifadhiwa. Angalia muundo wa faili.")
        except Exception as e:
            messages.error(request, f"Hitilafu ya kusoma faili: {e}")

        return redirect(f'{reverse("upload_form_students")}?form={selected_form}')

    # GET — show form. Ordered by insertion order (id), not alphabetically —
    # the Academic uploaded a specific roster order (e.g. matching the
    # school register) and expects to see it back exactly as uploaded.
    students = FormStudent.objects.filter(school=school, form=selected_form).order_by('id') if selected_form else FormStudent.objects.none()
    counts = {f: 0 for f in range(1, 7)}
    for row in FormStudent.objects.filter(school=school).values('form').annotate(cnt=Count('id')):
        if row['form'] in counts:
            counts[row['form']] = row['cnt']

    # Subject filter — when ?subject=<id> is provided, show only
    # students who study that subject (option subjects).
    subject_filter_id = request.GET.get('subject')
    # Subject is global (no school FK); use SchoolSubject for school-specific list.
    subjects_qs = Subject.objects.filter(schoolsubject__school=school).order_by('name').distinct() if school else Subject.objects.none()
    if selected_form and subject_filter_id:
        students = students.filter(subjects__id=subject_filter_id)

    # All subjects for the assign-subjects modal
    all_subjects = Subject.objects.filter(schoolsubject__school=school).order_by('name').distinct() if school else Subject.objects.none()

    return render(request, 'results/upload_form_students.html', {
        'selected_form': selected_form,
        'students': students,
        'form_counts': counts,
        'subjects_list': subjects_qs,
        'selected_subject': int(subject_filter_id) if subject_filter_id and subject_filter_id.isdigit() else None,
        'all_subjects': all_subjects,
    })


@academic_required
@require_POST
def assign_form_student_subjects(request, student_id):
    """Assign subjects to a FormStudent (option subjects like Physics, Agriculture)."""
    school = request.user.school
    student = get_object_or_404(FormStudent, id=student_id, school=school)
    subject_ids = request.POST.getlist('subjects')
    subjects = Subject.objects.filter(id__in=subject_ids)
    student.subjects.set(subjects)
    names = ', '.join(s.name for s in subjects) or 'Hakuna somo'
    messages.success(request, f"Masomo ya {student.full_name} yamewekwa: {names}")
    return redirect(f'{reverse("upload_form_students")}?form={student.form}')


@academic_required
@require_POST
def bulk_assign_form_student_subjects(request):
    """Bulk-assign subjects to multiple FormStudents at once.
    POST: student_ids=<id>&student_ids=<id>... subjects=<id>&subjects=<id>...
    action=add|set|remove
      add    — add selected subjects to each student (default)
      set    — replace all subjects with the selected ones
      remove — remove selected subjects from each student
    """
    school = request.user.school
    if not school:
        messages.error(request, "Hakuna shule.")
        return redirect('upload_form_students')

    student_ids = request.POST.getlist('student_ids')
    subject_ids = request.POST.getlist('subjects')
    action = request.POST.get('action', 'add')
    form_num = request.POST.get('form') or request.GET.get('form') or ''

    if not student_ids:
        messages.warning(request, "Chagua wanafunzi kwanza.")
        return redirect(f'{reverse("upload_form_students")}?form={form_num}')

    students = FormStudent.objects.filter(id__in=student_ids, school=school)
    subjects = Subject.objects.filter(id__in=subject_ids)

    count = 0
    for fs in students:
        if action == 'set':
            fs.subjects.set(subjects)
        elif action == 'remove':
            fs.subjects.remove(*subjects)
        else:  # add
            fs.subjects.add(*subjects)
        count += 1

    subj_names = ', '.join(s.name for s in subjects) or 'Hakuna somo'
    messages.success(request, f"Masomo yamewekwa kwa wanafunzi {count}: {subj_names}")
    return redirect(f'{reverse("upload_form_students")}?form={form_num}')


@academic_required
@require_POST
def delete_form_student(request, student_id):
    """Delete a single form student."""
    school = request.user.school
    student = get_object_or_404(FormStudent, id=student_id, school=school)
    form_num = student.form
    student.delete()
    messages.success(request, "Mwanafunzi ameondolewa.")
    return redirect(f'{reverse("upload_form_students")}?form={form_num}')


@academic_required
@require_POST
def delete_all_form_students(request, form_num):
    """Wipe the whole uploaded roster for one form — e.g. to start over
    with a corrected file, rather than deleting students one at a time."""
    school = request.user.school
    if form_num not in (1, 2, 3, 4, 5, 6):
        messages.error(request, "Form si sahihi.")
        return redirect('upload_form_students')
    deleted_count, _ = FormStudent.objects.filter(school=school, form=form_num).delete()
    if deleted_count:
        messages.success(request, f"Wanafunzi wote {deleted_count} wa Form {form_num} wameondolewa.")
    else:
        messages.info(request, f"Hakuna wanafunzi wa Form {form_num} wa kuondoa.")
    return redirect(f'{reverse("upload_form_students")}?form={form_num}')


@academic_required
def assign_teacher_form(request):
    """Academic Officer assigns teacher to form + subject."""
    school = request.user.school
    if not school:
        messages.error(request, "Hakuna shule iliyowekwa.")
        return redirect('home')

    if request.method == 'POST':
        teacher_id = request.POST.get('teacher_id')
        form_num = request.POST.get('form')
        subject_id = request.POST.get('subject_id')
        if teacher_id and form_num and subject_id:
            from .models import TeacherAccount
            teacher = TeacherAccount.objects.filter(id=teacher_id, school=school).first()
            subject = Subject.objects.filter(id=subject_id).first()
            if teacher and subject and int(form_num) in (1,2,3,4,5,6):
                TeacherFormAssignment.objects.get_or_create(
                    teacher=teacher, form=int(form_num), subject=subject, school=school,
                )
                messages.success(request, f"{teacher.full_name} ameassignwa Form {form_num} — {subject.name}")
            else:
                messages.error(request, "Taarifa hazijakamilika.")
        return redirect('assign_teacher_form')

    from .models import TeacherAccount
    teachers = TeacherAccount.objects.filter(school=school, role='TEACHER').order_by('full_name')
    subjects = Subject.objects.all().order_by('name')
    assignments = TeacherFormAssignment.objects.filter(school=school).select_related('teacher', 'subject').order_by('form', 'subject__name')

    return render(request, 'results/assign_teacher_form.html', {
        'teachers': teachers,
        'subjects': subjects,
        'assignments': assignments,
    })


def _generate_recommendations_for_teacher(teacher, form, school):
    """Generate automatic recommendations for a teacher based on subject performance."""
    from .models import ExamResult, Exam, ExamResult as ER
    from django.db.models import Avg, Count, Q

    recs = []
    assignment = TeacherFormAssignment.objects.filter(
        teacher=teacher, form=form, school=school
    ).first()
    if not assignment:
        return recs

    subject = assignment.subject

    # Find exams for this form at this school
    exams = Exam.objects.filter(school=school, form=form).order_by('-year', '-date')[:5]
    if not exams:
        recs.append({
            'type': 'info',
            'title': 'Hakuna mitihani bado',
            'detail': f'Hakuna mitihani ya Form {form} iliyopo. Tumia mitihani kuona matokeo ya somo la {subject.name}.',
        })
        return recs

    for exam in exams[:3]:
        results = ExamResult.objects.filter(exam=exam, subject=subject)
        if not results.exists():
            continue

        avg = results.aggregate(avg=Avg('score'))['avg'] or 0
        total = results.count()
        passed = results.filter(score__gte=40).count()
        pass_rate = (passed / total * 100) if total else 0
        failed = total - passed

        if pass_rate < 50:
            recs.append({
                'type': 'warning',
                'title': f'{exam.name} — Kiwango cha kufaulu ni chini ( {pass_rate:.0f}%)',
                'detail': f'Wanafunzi {failed} kati ya {total} wamefail {subject.name}. Fikiria: ongeza muda wa masomo, tumia mitihani ya ziada, au rekebisha mbinu za ufundishaji.',
            })
        elif avg < 45:
            recs.append({
                'type': 'info',
                'title': f'{exam.name} — Wastani ni {avg:.1f} (chini ya 45)',
                'detail': f'Wastani wa alama za {subject.name} ni {avg:.1f}. Wanafunzi wanahitaji msaada zaidi — ongeza mazoezi na ushauri wa karibu.',
            })
        elif pass_rate >= 80:
            recs.append({
                'type': 'success',
                'title': f'{exam.name} — Matokeo mazuri ({pass_rate:.0f}% wamefaulu)',
                'detail': f'Wanafunzi {passed} kati ya {total} wamefaulu {subject.name} kwa wastani wa {avg:.1f}. Endelea na mbinu ulizo nazo!',
            })
        else:
            recs.append({
                'type': 'info',
                'title': f'{exam.name} — Wastani {avg:.1f}, Kufaulu {pass_rate:.0f}%',
                'detail': f'Matokeo ni ya wastani. Fikiria kuongeza muda wa mazoezi na kuwasaidia wanafunzi walio chini ya 40.',
            })

    if not recs:
        recs.append({
            'type': 'info',
            'title': 'Hakuna data ya kutosha',
            'detail': f'Hakuna matokeo ya {subject.name} kwa Form {form} kwa ripoti ya mapendekezo.',
        })

    return recs


@academic_required
def teacher_performance_report(request, form_num):
    """Generate PDF report ya walimu wote wa form fulani — performance + recommendations."""
    school = request.user.school
    if not school:
        messages.error(request, "Hakuna shule iliyowekwa.")
        return redirect('home')

    if form_num not in (1,2,3,4,5,6):
        raise Http404("Form haijulikani")

    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.units import cm
    from reportlab.lib import colors

    assignments = TeacherFormAssignment.objects.filter(
        school=school, form=form_num
    ).select_related('teacher', 'subject').order_by('teacher__full_name', 'subject__name')

    form_labels = {1: 'I', 2: 'II', 3: 'III', 4: 'IV', 5: 'V', 6: 'VI'}
    form_label = form_labels.get(form_num, str(form_num))

    student_count = FormStudent.objects.filter(school=school, form=form_num).count()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
    content_w = A4[0] - 4*cm
    story = []

    ss = getSampleStyleSheet()
    title_s = ParagraphStyle('rpt_title', parent=ss['Title'], fontSize=16, alignment=TA_CENTER, textColor=colors.HexColor('#1A3C6E'), spaceAfter=6)
    subtitle_s = ParagraphStyle('rpt_sub', parent=ss['Normal'], fontSize=11, alignment=TA_CENTER, textColor=colors.HexColor('#555555'), spaceAfter=12)
    section_s = ParagraphStyle('rpt_sec', parent=ss['Heading2'], fontSize=12, textColor=colors.HexColor('#1A3C6E'), spaceBefore=16, spaceAfter=8)
    body_s = ParagraphStyle('rpt_body', parent=ss['Normal'], fontSize=10, leading=14, spaceAfter=4)
    rec_s = ParagraphStyle('rpt_rec', parent=ss['Normal'], fontSize=9, leading=13, leftIndent=12, spaceAfter=6)

    # Title
    story.append(Paragraph(f"{school.name}", title_s))
    story.append(Paragraph(f"Ripoti ya Ufaulu — Form {form_label}", subtitle_s))
    story.append(Paragraph(f"Wanafunzi: {student_count} | Walimu: {assignments.values('teacher').distinct().count()} | Masomo: {assignments.values('subject').distinct().count()}", subtitle_s))
    story.append(Spacer(1, 12))

    # Per-teacher sections
    teachers_seen = []
    for ta in assignments:
        teacher = ta.teacher
        if teacher.id in [t.id for t in teachers_seen]:
            continue
        teachers_seen.append(teacher)

        teacher_assignments = [a for a in assignments if a.teacher.id == teacher.id]
        teacher_subjects = ', '.join(a.subject.name for a in teacher_assignments)

        story.append(Paragraph(f"<b>{teacher.full_name or teacher.email}</b>", section_s))
        story.append(Paragraph(f"Masomo: {teacher_subjects}", body_s))

        # Performance per subject — batch-load all data once
        from django.db.models import Avg
        _form_exams = list(Exam.objects.filter(school=school, form=form_num).order_by('-year')[:5])
        _form_exam_ids = [e.id for e in _form_exams]
        _all_teacher_subject_ids = [a.subject.id for a in teacher_assignments]
        _perf_agg = ExamResult.objects.filter(
            exam_id__in=_form_exam_ids, subject_id__in=_all_teacher_subject_ids
        ).values('exam_id', 'subject_id').annotate(
            avg_score=Avg('score'),
            total=Count('id'),
            passed=Count('id', filter=Q(score__gte=40)),
        )
        _perf_map = {}
        for row in _perf_agg:
            _perf_map.setdefault(row['subject_id'], {})[row['exam_id']] = row

        for ta2 in teacher_assignments:
            subject = ta2.subject
            perf_data = [['Mtihani', 'Wastani', 'Kufaulu%', 'Idadi']]
            for ex in _form_exams:
                info = _perf_map.get(subject.id, {}).get(ex.id)
                if info and info['total']:
                    avg = info['avg_score'] or 0
                    total = info['total']
                    pr = round(info['passed'] / total * 100, 1) if total else 0
                    perf_data.append([ex.name, f'{avg:.1f}', f'{pr}%', str(total)])

            if len(perf_data) > 1:
                story.append(Paragraph(f"<b>{subject.name}</b>", body_s))
                t = Table(perf_data, colWidths=[content_w*0.35, content_w*0.2, content_w*0.2, content_w*0.15])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A3C6E')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                story.append(t)
                story.append(Spacer(1, 6))

        # Recommendations
        recs = _generate_recommendations_for_teacher(teacher, form_num, school)
        if recs:
            story.append(Paragraph("<b>Mapendekezo ya Kuimprove:</b>", body_s))
            for rec in recs:
                icon = {'warning': '⚠️', 'success': '✅', 'info': 'ℹ️'}.get(rec['type'], '•')
                story.append(Paragraph(f"{icon} <b>{rec['title']}</b>", rec_s))
                story.append(Paragraph(rec['detail'], rec_s))

        story.append(Spacer(1, 8))

    if not teachers_seen:
        story.append(Paragraph("Hakuna walimu walioassignwa kwa Form hii bado.", body_s))
        story.append(Paragraph("Tafadhali assign walimu kwanza (Assign Teacher → Form).", body_s))

    doc.build(story)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="Form_{form_label}_Teacher_Report.pdf"'
    return resp


# ══════════════════════════════════════════════════════════════════════════════
# BULK SCORESHEET UPLOAD — Academic officer uploads multiple PDFs/images
# (one per subject) and the system routes each to the correct subject.
# ══════════════════════════════════════════════════════════════════════════════

@academic_required
def bulk_scoresheet_upload(request, exam_id):
    """Academic officer uploads multiple scoresheets (PDFs/images) for
    different subjects at once.  Each file is OCR'd via a background
    Celery task, matched to students, and saved to the correct subject."""
    exam = _get_exam_or_404(exam_id, request.user)
    school = request.user.school

    # Get all subjects that have submissions for this exam
    subjects = list(
        Subject.objects.filter(submissions__exam=exam)
        .order_by('name')
    )
    if not subjects:
        # Fall back to subjects with results
        subjects = list(
            Subject.objects.filter(examresult__exam=exam).distinct().order_by('name')
        )

    # Get students in this exam for roster matching
    student_ids = list(
        ExamResult.objects.filter(exam=exam)
        .values_list('student_id', flat=True)
        .distinct()
    )
    roster_students = list(Student.objects.filter(id__in=student_ids))
    roster_ids = [s.id for s in roster_students]

    # Build per-subject status for the dashboard
    subject_status = []
    for sub in subjects:
        try:
            sub = Subject.objects.get(id=sub.id)
        except Subject.DoesNotExist:
            pass
        submission = SubjectSubmission.objects.filter(
            exam=exam, subject=sub
        ).first()
        has_results = ExamResult.objects.filter(
            exam=exam, subject=sub
        ).exists()
        subject_status.append({
            'subject': sub,
            'submission': submission,
            'has_results': has_results,
            'is_done': has_results or (submission and submission.status in ('SUBMITTED', 'APPROVED')),
        })

    if request.method == 'POST':
        import uuid as _uuid
        from django.core.files.storage import default_storage

        # Support both single and multiple file upload
        subject_id = request.POST.get('subject_id')
        files = request.FILES.getlist('scoresheet')
        subject_ids_raw = request.POST.getlist('subject_ids')

        # Multi-file upload: each file matched to a subject_id
        if subject_ids_raw and files and len(files) == len(subject_ids_raw):
            tasks_started = []
            for idx, (file, sid) in enumerate(zip(files, subject_ids_raw)):
                if not sid or not file:
                    continue
                try:
                    subject = Subject.objects.get(id=int(sid))
                except (Subject.DoesNotExist, ValueError):
                    continue
                ext = os.path.splitext(file.name)[1].lower() or '.pdf'
                storage_path = f"bulk_upload/{exam.id}/{subject.id}_{_uuid.uuid4().hex}{ext}"
                default_storage.save(storage_path, file)
                from .tasks import process_bulk_upload_task
                try:
                    task = process_bulk_upload_task.apply_async(
                        args=[storage_path, exam.id, subject.id, roster_ids],
                        kwargs={'preview_only': True},
                        queue='default',
                    )
                    tasks_started.append({'task_id': task.id, 'subject_id': subject.id, 'subject_name': subject.name})
                except Exception as celery_err:
                    # Celery broker may be down — run synchronously as fallback
                    logger.warning('Celery apply_async failed for subject %s, running synchronously: %s', subject.name, celery_err)
                    try:
                        result = process_bulk_upload_task(
                            storage_path, exam.id, subject.id, roster_ids, preview_only=True
                        )
                        if result and result.get('error'):
                            tasks_started.append({'error': result['error'], 'subject_id': subject.id, 'subject_name': subject.name})
                        else:
                            tasks_started.append({'task_id': None, 'subject_id': subject.id, 'subject_name': subject.name, 'sync_done': True, 'preview': result})
                    except Exception as sync_err:
                        logger.error('Synchronous bulk upload also failed for subject %s: %s', subject.name, sync_err)
                        tasks_started.append({'error': str(sync_err), 'subject_id': subject.id, 'subject_name': subject.name})
            return JsonResponse({'tasks': tasks_started})

        # Single file upload (legacy)
        if subject_id and files:
            file = files[0]
            subject = get_object_or_404(Subject, id=int(subject_id))
            ext = os.path.splitext(file.name)[1].lower() or '.pdf'
            storage_path = f"bulk_upload/{exam.id}/{subject.id}_{_uuid.uuid4().hex}{ext}"
            default_storage.save(storage_path, file)
            from .tasks import process_bulk_upload_task
            try:
                task = process_bulk_upload_task.apply_async(
                    args=[storage_path, exam.id, subject.id, roster_ids],
                    kwargs={'preview_only': True},
                    queue='default',
                )
                return JsonResponse({'task_id': task.id, 'subject_id': subject.id})
            except Exception as celery_err:
                logger.warning('Celery apply_async failed for subject %s, running synchronously: %s', subject_id, celery_err)
                try:
                    result = process_bulk_upload_task(
                        storage_path, exam.id, subject.id, roster_ids, preview_only=True
                    )
                    if result and result.get('error'):
                        return JsonResponse({'error': result['error']}, status=400)
                    return JsonResponse({'task_id': None, 'subject_id': subject.id, 'sync_done': True, 'preview': result})
                except Exception as sync_err:
                    logger.error('Synchronous bulk upload also failed for subject %s: %s', subject_id, sync_err)
                    return JsonResponse({'error': str(sync_err)}, status=500)

        return JsonResponse({'error': 'Tafadhali weka somo na upakie faili.'}, status=400)

    done_count = sum(1 for s in subject_status if s['is_done'])
    pending_count = len(subject_status) - done_count

    return render(request, 'results/bulk_scoresheet_upload.html', {
        'exam': exam,
        'subjects': subjects,
        'subject_status': subject_status,
        'subject_status_done': done_count,
        'subject_status_pending': pending_count,
        'student_count': len(roster_students),
    })


@academic_required
def ocr_health_check(request):
    """Check if OCR API keys are configured and working."""
    from .services.scoresheet_ocr_service import check_ocr_health
    health = check_ocr_health()
    return JsonResponse(health)


@academic_required
def bulk_upload_status(request, task_id):
    """Polled by the frontend every 2s after bulk upload kicks off."""
    from celery.result import AsyncResult
    result = AsyncResult(task_id)

    if not result.ready():
        return JsonResponse({'status': 'processing'})

    if result.failed():
        return JsonResponse({'error': 'Kuna hitilafu wakati wa kusoma scoresheet.'}, status=500)

    payload = result.result or {}
    if payload.get('error'):
        return JsonResponse({'error': payload['error']}, status=400)

    # Preview mode: task returned matched/unmatched without saving
    if payload.get('preview'):
        return JsonResponse({
            'status': 'preview',
            'preview': payload,
        })

    return JsonResponse({
        'status': 'done',
        'matched': payload.get('matched_count', 0),
        'unmatched_count': payload.get('unmatched_count', 0),
        'unmatched': payload.get('unmatched', []),
    })


@academic_required
@require_POST
def save_confirmed_scores(request, exam_id):
    """Save scores that the academic officer reviewed and confirmed.
    Called after the preview/OCR step — the frontend sends the final
    student_id + score pairs per subject."""
    import json as _json
    from django.utils import timezone as _tz
    from .services.upload_processing_service import recompute_processed_results_for_exam

    exam = _get_exam_or_404(exam_id, request.user)

    try:
        payload = _json.loads(request.body)
    except (ValueError, _json.JSONDecodeError):
        return JsonResponse({'error': 'Data ya ombi si sahihi.'}, status=400)

    subject_id = payload.get('subject_id')
    scores = payload.get('scores', [])  # [{student_id, score}, ...]

    if not subject_id or not scores:
        return JsonResponse({'error': 'Somo na alama lazima ziwe.'}, status=400)

    try:
        subject = Subject.objects.get(id=int(subject_id))
    except (Subject.DoesNotExist, ValueError):
        return JsonResponse({'error': 'Somo halipatikani.'}, status=400)

    # Build ExamResult entries from confirmed data
    exam_results = []
    for entry in scores:
        student_id = entry.get('student_id')
        score = entry.get('score')
        is_absent = entry.get('is_absent', False)
        if student_id is None:
            continue
        # For absent students, score can be None/0; for present, score must be valid
        if is_absent:
            score = None
        else:
            if score is None:
                continue
            try:
                score = int(score)
            except (TypeError, ValueError):
                continue
            if score < 0 or score > 100:
                continue
        exam_results.append(ExamResult(
            exam=exam, student_id=int(student_id), subject=subject,
            score=score if not is_absent else None,
            is_absent=bool(is_absent),
        ))

    if exam_results:
        ExamResult.objects.bulk_create(
            exam_results,
            update_conflicts=True,
            unique_fields=['exam', 'student', 'subject'],
            update_fields=['score', 'is_absent'],
        )

    # Mark SubjectSubmission as SUBMITTED + APPROVED
    SubjectSubmission.objects.update_or_create(
        exam=exam, subject=subject,
        defaults={
            'status': SubjectSubmission.STATUS_APPROVED,
            'method': 'UPLOAD',
            'submitted_by': 'Academic Officer (Bulk Upload)',
            'submitted_at': _tz.now(),
            'approved_by': 'Academic Officer (Bulk Upload)',
            'approved_at': _tz.now(),
            'student_count': len(exam_results),
        },
    )

    # Recompute processed results
    recompute_processed_results_for_exam(exam)

    return JsonResponse({
        'status': 'done',
        'saved_count': len(exam_results),
    })
