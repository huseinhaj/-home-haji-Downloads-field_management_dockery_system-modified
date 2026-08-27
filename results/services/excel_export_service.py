"""
excel_export_service.py — Professional multi-sheet Excel with Tanzania flag colours.

Three sheets:
  Sheet 1  Full Results / Matokeo Kamili — per-student scores + totals
  Sheet 2  Summary / Muhtasari — division breakdown, subject stats, top 5
  Sheet 3  Subject Analysis / Somo kwa Somo — per-subject ranking + grades

Language: English for secondary schools, Kiswahili for primary schools.
School name: "KWADELO SECONDARY SCHOOL" not just "KWADELO".
Colours: Tanzania flag (green #1EB53A, yellow #FCD116, black, blue #00A3DD).
"""

from __future__ import annotations

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .export_data import get_exam_export_payload
from .report_helpers import (
    TZ_BLUE,
    TZ_DARK_GREY,
    TZ_GOLD,
    TZ_GREEN,
    TZ_LIGHT_GREY,
    TZ_WHITE,
    get_full_school_name,
    get_report_label,
    get_report_language,
    get_section_title,
    get_school_type_for_exam,
)


# ── Color palette — Tanzania flag (FF prefix for openpyxl aRGB) ────────────
GREEN = "FF" + TZ_GREEN
GOLD  = "FF" + TZ_GOLD
BLUE  = "FF" + TZ_BLUE
WHITE = "FF" + TZ_WHITE
LIGHT_GREY = "FF" + TZ_LIGHT_GREY
DARK_GREY  = "FF" + TZ_DARK_GREY


def _grade_thresholds(form):
    """Return (score-thresholds, grade-ranges) for the exam's NECTA scale.

    Form 5/6 (ACSEE) grades on a wider scale than Form 1/2/3/4 (CSEE/FTNA
    share the same 5-band A/B/C/D/F scale — verified against real NECTA
    CSEE result slips, which never show B+/C+):
        CSEE:  A 75+ | B 65+ | C 45+ | D 30+ | F <30
        FTNA:  A 75+ | B 65+ | C 45+ | D 30+ | F <30
        ACSEE: A 80+ | B 70+ | C 60+ | D 50+ | E 40+ | S 35+ | F <35
    """
    if form == 2:
        return [75, 65, 45, 30], [('A', '75-100'), ('B', '65-74'), ('C', '45-64'), ('D', '30-44'), ('F', '0-29')]
    if form in (5, 6):
        return [80, 70, 60, 50, 40, 35], [('A', '80-100'), ('B', '70-79'), ('C', '60-69'), ('D', '50-59'), ('E', '40-49'), ('S', '35-39'), ('F', '0-34')]
    return [75, 65, 45, 30], [('A', '75-100'), ('B', '65-74'), ('C', '45-64'), ('D', '30-44'), ('F', '0-29')]


# Grade colour per letter — shared by score cells and the grading-key legend.
# Must stay in sync with utils.get_grade_for_form / get_grade_points.
_FILL_BY_LETTER = {
    'A':  ("FFC6F4D6", "FF145A32"),
    'B':  ("FFD5F5E3", "FF1E8449"),
    'C':  ("FFFFF9C4", "FF7D6608"),
    'D':  ("FFFDEBD0", "FF784212"),
    'E':  ("FFF0B27A", "FF9C640C"),
    'S':  ("FFF9E79F", "FFB9770B"),
    'F':  ("FFFADBD8", "FF922B21"),
}


def _score_fill(score, form=4) -> tuple[str | None, str | None]:
    if not isinstance(score, (int, float)):
        return None, None
    thresholds, grade_ranges = _grade_thresholds(form)
    for i, t in enumerate(thresholds):
        if score >= t:
            return _FILL_BY_LETTER.get(grade_ranges[i][0], (None, None))
    return _FILL_BY_LETTER.get(grade_ranges[-1][0], (None, None))


def _score_grade(score, form=4) -> str:
    if not isinstance(score, (int, float)):
        return '-'
    thresholds, grade_ranges = _grade_thresholds(form)
    for i, t in enumerate(thresholds):
        if score >= t:
            return grade_ranges[i][0]
    return grade_ranges[-1][0]


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _make_border(color: str = "FFCCCCCC") -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _green_header_style(cell, *, bold: bool = True):
    cell.font = Font(color=WHITE, bold=bold, name='Calibri', size=10)
    cell.fill = _make_fill(GREEN)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = _make_border(GOLD)


def _gold_header_style(cell):
    cell.font = Font(color=GREEN, bold=True, name='Calibri', size=9)
    cell.fill = _make_fill(GOLD)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _make_border(GREEN)


def _set_auto_width(ws, min_width: int = 8, max_width: int = 35):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                try:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


# ── Sheet 1: Full Results / Matokeo Kamili ──────────────────────────────────
def _build_sheet_matokeo(wb: openpyxl.Workbook, exam, payload: dict):
    lang = get_report_language(exam)
    school_disp = get_full_school_name(exam)

    ws = wb.active
    if lang == 'sw':
        ws.title = "Matokeo Kamili"
    else:
        ws.title = "Full Results"

    subjects = payload['subjects']
    results = payload['processed_results']
    score_lookup = payload['score_lookup']
    absent_lookup = payload['absent_lookup']
    student_subjects = payload['student_subjects']

    total_cols = 3 + len(subjects) + 4  # POS + NAME + SEX + subjects + TOTAL + AVG + DIV + PTS

    # Title rows
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    if lang == 'sw':
        title_cell.value = f"{school_disp} — MATOKEO YA {exam.get_exam_type_display().upper()} {exam.year}"
    else:
        title_cell.value = f"{school_disp} — {exam.get_exam_type_display().upper()} EXAMINATION {exam.year}"
    title_cell.font = Font(bold=True, size=14, color=WHITE, name='Calibri')
    title_cell.fill = _make_fill(GREEN)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    sub_cell = ws.cell(row=2, column=1)
    if lang == 'sw':
        sub_cell.value = f"Fomu {exam.form}   |   {exam.name}   |   Mwaka {exam.year}"
    else:
        sub_cell.value = f"Form {exam.form}   |   {exam.name}   |   Year {exam.year}"
    sub_cell.font = Font(bold=False, size=10, color=GOLD, name='Calibri')
    sub_cell.fill = _make_fill(GREEN)
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # Header row
    header_row = 3
    if lang == 'sw':
        headers = ["POS", "JINA", "JINSIA"] + [s.name.upper() for s in subjects] + ["JUMLA", "WASTANI", "DARAJA", "POINTI"]
    else:
        headers = ["POS", "NAME", "SEX"] + [s.name.upper() for s in subjects] + ["TOTAL", "AVG", "DIV.", "PTS"]
    for col_idx, title in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        _green_header_style(cell)
    ws.row_dimensions[header_row].height = 22

    # Freeze panes
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Data rows
    subject_start_col = 4  # columns 1-3 = POS, NAME/SEX
    for row_idx, result in enumerate(results):
        data_row = header_row + 1 + row_idx
        student = result.student
        full_name = ' '.join(
            part for part in [student.first_name, student.middle_name or '', student.last_name] if part
        ).strip()

        ws.cell(row=data_row, column=1, value=result.position)
        ws.cell(row=data_row, column=2, value=full_name)
        ws.cell(row=data_row, column=3, value=student.gender)

        row_fill = _make_fill(LIGHT_GREY) if row_idx % 2 == 0 else None

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.font = Font(name='Calibri', size=9)
            cell.border = _make_border()
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if row_fill and col_idx not in (2,):
                cell.fill = row_fill
            if col_idx == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # Subject scores — only show subjects the student is enrolled in
        enrolled_ids = student_subjects.get(student.id, set())
        for sub_col_offset, subject in enumerate(subjects):
            col_idx = subject_start_col + sub_col_offset
            cell = ws.cell(row=data_row, column=col_idx)
            if subject.id not in enrolled_ids:
                # Student does not study this subject — leave blank
                cell.value = ''
            elif (student.id, subject.id) in absent_lookup:
                # Student studies this subject but was absent
                cell.value = 'X'
                cell.font = Font(color='FF475569', bold=True, name='Calibri', size=9)
            else:
                score = score_lookup.get((student.id, subject.id))
                if score is not None:
                    cell.value = score
                    fill_hex, font_hex = _score_fill(score, exam.form)
                    if fill_hex:
                        cell.fill = _make_fill(fill_hex)
                    if font_hex:
                        cell.font = Font(color=font_hex, bold=True, name='Calibri', size=9)
                else:
                    cell.value = '-'

        # Totals columns
        total_col = subject_start_col + len(subjects)
        avg_col = total_col + 1
        div_col = avg_col + 1
        pts_col = div_col + 1

        total_cell = ws.cell(row=data_row, column=total_col, value=result.total_score)
        total_cell.font = Font(bold=True, name='Calibri', size=9, color=GREEN)

        avg_cell = ws.cell(row=data_row, column=avg_col, value=float(result.average_score))
        avg_cell.number_format = '0.00'
        avg_cell.font = Font(bold=True, name='Calibri', size=9)

        div_cell = ws.cell(row=data_row, column=div_col, value=result.division)
        if result.division == 'I':
            div_cell.fill = _make_fill("FFC6F4D6")
            div_cell.font = Font(bold=True, name='Calibri', size=9, color="FF145A32")
        elif result.division in ('II', 'III'):
            div_cell.fill = _make_fill("FFFFF9C4")
            div_cell.font = Font(bold=True, name='Calibri', size=9, color="FF7D6608")
        elif result.division in ('IV', '0'):
            div_cell.fill = _make_fill("FFFADBD8")
            div_cell.font = Font(bold=True, name='Calibri', size=9, color="FF922B21")

        ws.cell(row=data_row, column=pts_col, value=result.points)

    # Grade legend — ranges follow the exam's NECTA scale (FTNA vs CSEE)
    last_data_row = header_row + len(results) + 1
    legend_row = last_data_row + 2
    _, grade_ranges = _grade_thresholds(exam.form)
    legend_labels = [
        (f"{g}  ({rng})", *_FILL_BY_LETTER.get(g, ("FFFFFFFF", "FF000000")))
        for g, rng in grade_ranges
    ]
    if lang == 'sw':
        ws.cell(row=legend_row, column=1, value="UFUNGUO WA DARAJA:")
    else:
        ws.cell(row=legend_row, column=1, value="GRADING KEY:")
    ws.cell(row=legend_row, column=1).font = Font(bold=True, name='Calibri', size=8, color=GREEN)
    for i, (label, bg, fg) in enumerate(legend_labels):
        cell = ws.cell(row=legend_row, column=2 + i, value=label)
        cell.fill = _make_fill(bg)
        cell.font = Font(bold=True, name='Calibri', size=8, color=fg)
        cell.alignment = Alignment(horizontal='center')
        cell.border = _make_border()

    _set_auto_width(ws)
    ws.column_dimensions['B'].width = 28


# ── Sheet 2: Summary / Muhtasari ────────────────────────────────────────────
def _build_sheet_muhtasari(wb: openpyxl.Workbook, exam, payload: dict):
    lang = get_report_language(exam)
    school_disp = get_full_school_name(exam)

    if lang == 'sw':
        ws = wb.create_sheet(title="Muhtasari")
    else:
        ws = wb.create_sheet(title="Summary")

    subjects = payload['subjects']
    results = payload['processed_results']
    score_lookup = payload['score_lookup']

    def _write_section_header(row, en, sw):
        title = sw if lang == 'sw' else en
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=10, color=WHITE, name='Calibri')
        cell.fill = _make_fill(GREEN)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = _make_border(GOLD)
        ws.row_dimensions[row].height = 18

    def _write_kv(row, key, val):
        k_cell = ws.cell(row=row, column=1, value=key)
        k_cell.font = Font(bold=True, name='Calibri', size=9, color=DARK_GREY)
        k_cell.fill = _make_fill(LIGHT_GREY)
        k_cell.border = _make_border()
        v_cell = ws.cell(row=row, column=2, value=val)
        v_cell.font = Font(name='Calibri', size=9)
        v_cell.border = _make_border()

    cur_row = 1

    # Exam details
    _write_section_header(cur_row, "EXAM INFORMATION", "MAELEZO YA MTIHANI")
    cur_row += 1
    _write_kv(cur_row, "Exam Name" if lang == 'en' else "Jina la Mtihani", exam.name)
    cur_row += 1
    _write_kv(cur_row, "Exam Type" if lang == 'en' else "Aina ya Mtihani", exam.get_exam_type_display())
    cur_row += 1
    _write_kv(cur_row, "Year" if lang == 'en' else "Mwaka", exam.year)
    cur_row += 1
    _write_kv(cur_row, f"Form {exam.form}" if lang == 'en' else f"Fomu {exam.form}", f"Form {exam.form}")
    cur_row += 1
    if exam.school_name or exam.school:
        _write_kv(cur_row, "School" if lang == 'en' else "Shule", school_disp)
        cur_row += 1
    _write_kv(cur_row, "Total Students" if lang == 'en' else "Jumla ya Wanafunzi", len(results))
    cur_row += 2

    # Division breakdown
    _write_section_header(cur_row, "DIVISION BREAKDOWN", "MGAWANYO WA DARAJA")
    cur_row += 1
    if lang == 'sw':
        div_headers = ["DARAJA", "IDADI", "ASILIMIA"]
    else:
        div_headers = ["DIVISION", "COUNT", "PERCENTAGE"]
    for col_idx, h in enumerate(div_headers, 1):
        cell = ws.cell(row=cur_row, column=col_idx, value=h)
        _gold_header_style(cell)
    cur_row += 1

    div_counts = {'I': 0, 'II': 0, 'III': 0, 'IV': 0, '0': 0}
    for r in results:
        div_counts[r.division] = div_counts.get(r.division, 0) + 1

    total_students = len(results) or 1
    for div in ['I', 'II', 'III', 'IV', '0']:
        count = div_counts.get(div, 0)
        pct = round(count / total_students * 100, 1)
        if lang == 'sw':
            ws.cell(row=cur_row, column=1, value=f"Daraja {div}")
        else:
            div_names = {'I': 'Division I', 'II': 'Division II', 'III': 'Division III',
                         'IV': 'Division IV', '0': 'Fail (0)'}
            ws.cell(row=cur_row, column=1, value=div_names[div])
        ws.cell(row=cur_row, column=1).border = _make_border()
        ws.cell(row=cur_row, column=2, value=count).border = _make_border()
        ws.cell(row=cur_row, column=3, value=f"{pct}%").border = _make_border()
        for col_idx in range(1, 4):
            ws.cell(row=cur_row, column=col_idx).font = Font(name='Calibri', size=9)
            ws.cell(row=cur_row, column=col_idx).alignment = Alignment(horizontal='center')
        cur_row += 1
    cur_row += 1

    # Per-subject statistics
    _write_section_header(cur_row, "SUBJECT STATISTICS", "TAKWIMU ZA MASOMO")
    cur_row += 1
    if lang == 'sw':
        sub_headers = ["SOMO", "WASTANI", "JUU ZAIDI", "CHINI ZAIDI", "ASILIMIA KUFAULU"]
    else:
        sub_headers = ["SUBJECT", "AVERAGE", "HIGHEST", "LOWEST", "PASS %"]
    for col_idx, h in enumerate(sub_headers, 1):
        cell = ws.cell(row=cur_row, column=col_idx, value=h)
        _gold_header_style(cell)
    cur_row += 1

    for subject in subjects:
        scores = [
            score_lookup[(r.student_id, subject.id)]
            for r in results
            if (r.student_id, subject.id) in score_lookup
        ]
        if not scores:
            continue
        avg = round(sum(scores) / len(scores), 1)
        highest = max(scores)
        lowest = min(scores)
        passing = sum(1 for s in scores if s >= 40)
        pass_pct = round(passing / len(scores) * 100, 1)

        ws.cell(row=cur_row, column=1, value=subject.name)
        ws.cell(row=cur_row, column=2, value=avg)
        ws.cell(row=cur_row, column=3, value=highest)
        ws.cell(row=cur_row, column=4, value=lowest)
        ws.cell(row=cur_row, column=5, value=f"{pass_pct}%")

        for col_idx in range(1, 6):
            cell = ws.cell(row=cur_row, column=col_idx)
            cell.font = Font(name='Calibri', size=9)
            cell.border = _make_border()
            cell.alignment = Alignment(horizontal='center')
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal='left')
        cur_row += 1
    cur_row += 1

    # Top 5 students
    _write_section_header(cur_row, "TOP 5 PERFORMERS", "WANAFUNZI BORA 5")
    cur_row += 1
    if lang == 'sw':
        top_headers = ["NAFASI", "JINA", "JUMLA", "WASTANI", "DARAJA"]
    else:
        top_headers = ["POS.", "NAME", "TOTAL", "AVERAGE", "DIV."]
    for col_idx, h in enumerate(top_headers, 1):
        cell = ws.cell(row=cur_row, column=col_idx, value=h)
        _gold_header_style(cell)
    cur_row += 1

    for result in results[:5]:
        student = result.student
        full_name = ' '.join(
            part for part in [student.first_name, student.middle_name or '', student.last_name] if part
        ).strip()
        ws.cell(row=cur_row, column=1, value=result.position)
        ws.cell(row=cur_row, column=2, value=full_name)
        ws.cell(row=cur_row, column=3, value=result.total_score)
        ws.cell(row=cur_row, column=4, value=float(result.average_score))
        ws.cell(row=cur_row, column=5, value=result.division)
        for col_idx in range(1, 6):
            cell = ws.cell(row=cur_row, column=col_idx)
            cell.font = Font(name='Calibri', size=9)
            cell.border = _make_border()
            cell.alignment = Alignment(horizontal='center')
        ws.cell(row=cur_row, column=2).alignment = Alignment(horizontal='left')
        cur_row += 1

    _set_auto_width(ws)
    ws.column_dimensions['B'].width = 28


# ── Sheet 3: Subject Analysis / Somo kwa Somo ───────────────────────────────
def _build_sheet_somo_kwa_somo(wb: openpyxl.Workbook, exam, payload: dict):
    lang = get_report_language(exam)

    if lang == 'sw':
        ws = wb.create_sheet(title="Somo kwa Somo")
    else:
        ws = wb.create_sheet(title="Subject Analysis")

    subjects = payload['subjects']
    results = payload['processed_results']
    score_lookup = payload['score_lookup']

    cur_row = 1

    # Title
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
    title_val = "UCHAMBUZI WA MASOMO" if lang == 'sw' else "SUBJECT ANALYSIS"
    title_cell = ws.cell(row=cur_row, column=1, value=title_val)
    title_cell.font = Font(bold=True, size=12, color=WHITE, name='Calibri')
    title_cell.fill = _make_fill(GREEN)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[cur_row].height = 22
    cur_row += 2

    for subject in subjects:
        # Subject header
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=4)
        subj_cell = ws.cell(row=cur_row, column=1, value=subject.name.upper())
        subj_cell.font = Font(bold=True, size=10, color=WHITE, name='Calibri')
        subj_cell.fill = _make_fill(GREEN)
        subj_cell.alignment = Alignment(horizontal='left', vertical='center')
        subj_cell.border = _make_border(GOLD)
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        # Column headers
        if lang == 'sw':
            col_headers = ["NAFASI", "JINA", "ALAMA", "DARAJA"]
        else:
            col_headers = ["POS.", "NAME", "SCORE", "GRADE"]
        for col_idx, h in enumerate(col_headers, 1):
            cell = ws.cell(row=cur_row, column=col_idx, value=h)
            _gold_header_style(cell)
        cur_row += 1

        # Collect and sort by score
        subject_rows = []
        for result in results:
            score = score_lookup.get((result.student_id, subject.id))
            if score is None:
                continue
            student = result.student
            full_name = ' '.join(
                part for part in [student.first_name, student.middle_name or '', student.last_name] if part
            ).strip()
            subject_rows.append({'name': full_name, 'score': score})

        subject_rows.sort(key=lambda r: r['score'], reverse=True)

        for rank, row in enumerate(subject_rows, 1):
            grade = _score_grade(row['score'], exam.form)
            fill_hex, font_hex = _score_fill(row['score'], exam.form)

            ws.cell(row=cur_row, column=1, value=rank)
            ws.cell(row=cur_row, column=2, value=row['name'])
            score_cell = ws.cell(row=cur_row, column=3, value=row['score'])
            grade_cell = ws.cell(row=cur_row, column=4, value=grade)

            if fill_hex:
                score_cell.fill = _make_fill(fill_hex)
                grade_cell.fill = _make_fill(fill_hex)
            if font_hex:
                score_cell.font = Font(bold=True, name='Calibri', size=9, color=font_hex)
                grade_cell.font = Font(bold=True, name='Calibri', size=9, color=font_hex)

            for col_idx in range(1, 5):
                cell = ws.cell(row=cur_row, column=col_idx)
                cell.border = _make_border()
                cell.alignment = Alignment(horizontal='center')
                if col_idx != 3 and col_idx != 4:
                    if not cell.font or not cell.font.bold:
                        cell.font = Font(name='Calibri', size=9)
            ws.cell(row=cur_row, column=2).alignment = Alignment(horizontal='left')
            cur_row += 1

        cur_row += 1  # gap between subjects

    _set_auto_width(ws)
    ws.column_dimensions['B'].width = 28


# ── Public API ───────────────────────────────────────────────────────────────
def generate_professional_excel_response(exam) -> HttpResponse:
    """Generate a 3-sheet professional Excel for the given exam."""
    payload = get_exam_export_payload(exam)

    wb = openpyxl.Workbook()
    _build_sheet_matokeo(wb, exam, payload)
    _build_sheet_muhtasari(wb, exam, payload)
    _build_sheet_somo_kwa_somo(wb, exam, payload)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = exam.name.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename=\"{safe_name}_{exam.year}_Results.xlsx\"'
    wb.save(response)
    return response


def generate_results_excel_response(exam) -> HttpResponse:
    """Legacy alias — now returns the professional multi-sheet Excel."""
    return generate_professional_excel_response(exam)
