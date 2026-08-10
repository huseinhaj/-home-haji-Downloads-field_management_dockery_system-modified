#!/usr/bin/env python3
"""
export_tanzania_geodata.py
===========================
Tengeneza Excel kamili ya Tanzania: Mikoa yote, Wilaya zote (kwa kila mkoa),
na Shule zote (kwa kila wilaya) — kutoka fixture files.

Matumizi:
    python3 export_tanzania_geodata.py

Matokeo:
    Tanzania_Mikoa_Wilaya_Shule.xlsx  (sheets: Mikoa, Wilaya, Shule)
"""

import json
import os
import sys
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Rangi za bendera ya Tanzania ─────────────────────────────────────────────
GREEN = "FF1EB53A"
GOLD  = "FFFCD116"
BLUE  = "FF00A3DD"
BLACK = "FF000000"
WHITE = "FFFFFFFF"
LIGHT_GREY = "FFF2F2F2"
DARK_GREY  = "FF333333"

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _border(color="FFCCCCCC"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def _green_header(cell, size=10):
    cell.font = Font(color=WHITE, bold=True, name='Calibri', size=size)
    cell.fill = _fill(GREEN)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = _border(GOLD)

def _gold_header(cell):
    cell.font = Font(color=GREEN, bold=True, name='Calibri', size=9)
    cell.fill = _fill(GOLD)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _border(GREEN)

def _auto_width(ws, min_w=8, max_w=45):
    for col in ws.columns:
        max_len = min_w
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[letter].width = min(max_len + 3, max_w)

# ── Load fixtures ─────────────────────────────────────────────────────────────
def load_fixture(path):
    with open(path, encoding='utf-8') as f:
        return json.load(f)

def load_data():
    regions   = load_fixture(os.path.join(BASE_DIR, 'regions.json'))
    districts = load_fixture(os.path.join(BASE_DIR, 'districts.json'))

    # Schools: main fixtures + missing_schools (ziadane)
    schools = load_fixture(os.path.join(BASE_DIR, 'schools.json'))
    try:
        schools += load_fixture(os.path.join(BASE_DIR, 'field_app/fixtures/missing_schools.json'))
    except FileNotFoundError:
        pass

    # De-duplicate schools by pk (missing_schools may overlap)
    seen = set()
    unique_schools = []
    for s in schools:
        pk = s.get('pk')
        if pk in seen:
            continue
        seen.add(pk)
        unique_schools.append(s)
    schools = unique_schools

    region_by_id = {r['pk']: r['fields']['name'] for r in regions}
    district_by_id = {d['pk']: d['fields'] for d in districts}

    # Organize districts by region
    districts_by_region = OrderedDict()
    for d in sorted(districts, key=lambda x: (region_by_id.get(x['fields']['region'], ''), x['fields']['name'])):
        rid = d['fields']['region']
        districts_by_region.setdefault(rid, []).append(d)

    # Organize schools by district
    schools_by_district = OrderedDict()
    for s in schools:
        did = s['fields'].get('district')
        schools_by_district.setdefault(did, []).append(s)

    return regions, districts, schools, region_by_id, district_by_id, districts_by_region, schools_by_district


def build_workbook(data):
    regions, districts, schools, region_by_id, district_by_id, districts_by_region, schools_by_district = data

    wb = openpyxl.Workbook()

    # ═══ SHEET 1: MIKOA ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Mikoa"
    headers = ["NAFASI", "MKOA", "IDADI YA WILAYA", "IDADI YA SHULE"]
    for col, h in enumerate(headers, 1):
        _green_header(ws.cell(row=1, column=col, value=h))
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    row = 2
    for i, r in enumerate(sorted(regions, key=lambda x: x['fields']['name']), 1):
        rid = r['pk']
        n_districts = len(districts_by_region.get(rid, []))
        n_schools = sum(len(schools_by_district.get(d['pk'], [])) for d in districts_by_region.get(rid, []))
        vals = [i, r['fields']['name'], n_districts, n_schools]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = Font(name='Calibri', size=10)
            cell.border = _border()
            if col != 2:
                cell.alignment = Alignment(horizontal='center')
            if row % 2 == 0:
                cell.fill = _fill(LIGHT_GREY)
        ws.cell(row=row, column=2).font = Font(name='Calibri', size=10, bold=True)
        row += 1
    _auto_width(ws)

    # ═══ SHEET 2: WILAYA ═════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Wilaya")
    headers2 = ["MKOA", "WILAYA", "IDADI YA SHULE"]
    for col, h in enumerate(headers2, 1):
        _green_header(ws2.cell(row=1, column=col, value=h))
    ws2.row_dimensions[1].height = 22
    ws2.freeze_panes = "A2"

    row = 2
    for rid in sorted(districts_by_region.keys(), key=lambda x: region_by_id.get(x, '')):
        region_name = region_by_id.get(rid, '?')
        for d in sorted(districts_by_region[rid], key=lambda x: x['fields']['name']):
            n_schools = len(schools_by_district.get(d['pk'], []))
            vals = [region_name, d['fields']['name'], n_schools]
            for col, v in enumerate(vals, 1):
                cell = ws2.cell(row=row, column=col, value=v)
                cell.font = Font(name='Calibri', size=10)
                cell.border = _border()
                if col != 2:
                    cell.alignment = Alignment(horizontal='center')
                if row % 2 == 0:
                    cell.fill = _fill(LIGHT_GREY)
            ws2.cell(row=row, column=2).font = Font(name='Calibri', size=10, bold=True)
            row += 1
    _auto_width(ws2)

    # ═══ SHEET 3: SHULE ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Shule")
    headers3 = ["MKOA", "WILAYA", "JINA LA SHULE", "AINA", "CODE", "UMILIKI", "MKUU WA SHULE", "SIMU"]
    for col, h in enumerate(headers3, 1):
        _green_header(ws3.cell(row=1, column=col, value=h))
    ws3.row_dimensions[1].height = 22
    ws3.freeze_panes = "A2"

    LEVEL_DISPLAY = {'Primary': 'Msingi', 'Secondary': 'Sekondari'}
    OWN_DISPLAY = {'government': 'Serikali', 'private': 'Binafsi'}

    row = 2
    for did in sorted(schools_by_district.keys()):
        d_fields = district_by_id.get(did, {})
        district_name = d_fields.get('name', '?')
        region_name = region_by_id.get(d_fields.get('region'), '?')
        for s in sorted(schools_by_district[did], key=lambda x: x['fields'].get('name', '')):
            f = s['fields']
            vals = [
                region_name,
                district_name,
                f.get('name', ''),
                LEVEL_DISPLAY.get(f.get('level', ''), f.get('level', '')),
                f.get('school_code', ''),
                OWN_DISPLAY.get(f.get('ownership', ''), ''),
                f.get('head_name', ''),
                f.get('head_phone', ''),
            ]
            for col, v in enumerate(vals, 1):
                cell = ws3.cell(row=row, column=col, value=v)
                cell.font = Font(name='Calibri', size=9)
                cell.border = _border()
                if row % 2 == 0:
                    cell.fill = _fill(LIGHT_GREY)
            ws3.cell(row=row, column=3).font = Font(name='Calibri', size=9, bold=True)
            row += 1
    _auto_width(ws3)
    ws3.column_dimensions['C'].width = 45

    return wb


def main():
    data = load_data()
    regions, districts, schools, *_ = data
    print(f"Mikoa: {len(regions)} | Wilaya: {len(districts)} | Shule: {len(schools)}")

    wb = build_workbook(data)
    out = os.path.join(BASE_DIR, "Tanzania_Mikoa_Wilaya_Shule.xlsx")
    wb.save(out)
    print(f"✅ Excel imetengenezwa: {out}")
    print("   Sheets: Mikoa | Wilaya | Shule")


if __name__ == '__main__':
    sys.exit(main())
